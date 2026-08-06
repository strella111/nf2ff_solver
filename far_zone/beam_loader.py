# -*- coding: utf-8 -*-
"""Загрузка результатов режима «Измерение лучей АФАР» (Beam№*.xlsx + scan_params.json).

Выделено из excel_module.py основного проекта — только функция-загрузчик.
"""

import os
import json

import numpy as np
from loguru import logger
from openpyxl import load_workbook
from typing import Optional


class LoadCancelled(Exception):
    """Загрузка прервана пользователем — это не ошибка данных."""


class _Progress:
    """Ход загрузки: доля выполнения 0..1 + текст, с проверкой отмены.

    on_progress(frac, text) — куда сообщать; frac=None означает «идёт длинная
    операция, доля неизвестна» (разбор книги Excel целиком).
    is_cancelled() — True, если пользователь нажал «Отмена».
    part() выделяет вложенный отчёт под подзадачу, которая занимает часть
    общего диапазона (например, один файл луча из сорока).
    """

    def __init__(self, on_progress=None, is_cancelled=None, lo=0.0, hi=1.0):
        self._on_progress = on_progress
        self._is_cancelled = is_cancelled
        self._lo = lo
        self._hi = hi

    def check(self):
        """Прервать загрузку, если пользователь нажал «Отмена»."""
        if self._is_cancelled is not None and self._is_cancelled():
            raise LoadCancelled()

    def emit(self, frac, text):
        """Сообщить долю выполнения внутри своего диапазона (frac ∈ 0..1)."""
        self.check()
        if self._on_progress is None:
            return
        if frac is None:
            self._on_progress(None, text)
            return
        frac = min(max(float(frac), 0.0), 1.0)
        self._on_progress(self._lo + (self._hi - self._lo) * frac, text)

    def busy(self, text):
        """Длинная операция без измеримой доли (открытие книги Excel)."""
        self.emit(None, text)

    def part(self, frac_from, frac_to):
        """Вложенный отчёт для подзадачи, занимающей часть диапазона."""
        span = self._hi - self._lo
        return _Progress(self._on_progress, self._is_cancelled,
                         self._lo + span * frac_from, self._lo + span * frac_to)


def load_beam_pattern_results(save_dir: str, on_progress=None,
                              is_cancelled=None) -> Optional[dict]:
    """
    Загружает результаты измерения лучей из Excel файлов для досканирования

    Args:
        save_dir: Путь к папке с результатами (base_dir/luchi/{дата})
        on_progress: callback(frac, text) — ход загрузки (frac ∈ 0..1 или None)
        is_cancelled: callback() -> bool — пользователь нажал «Отмена»

    Raises:
        LoadCancelled — загрузка прервана пользователем.

    Returns:
        dict: {
            'beams': [список лучей],
            'freq_list': [список частот],
            'data': {beam_num: {freq: {'x': [...], 'y': [...], 'amp': [[...]], 'phase': [[...]]}}},
            'x_list': [список координат X],
            'y_list': [список координат Y],
            'step_x': шаг по X,
            'step_y': шаг по Y
        } или None при ошибке
    """
    root = _Progress(on_progress, is_cancelled)
    try:
        if not os.path.exists(save_dir):
            logger.error(f"Папка не найдена: {save_dir}")
            return None

        root.emit(0.0, 'Поиск файлов измерения…')
        params_file = os.path.join(save_dir, 'scan_params.json')
        loaded_params = None
        if os.path.exists(params_file):
            try:
                with open(params_file, 'r', encoding='utf-8') as f:
                    loaded_params = json.load(f)
                logger.info(f"Загружены параметры сканирования из {params_file}")
            except Exception as e:
                logger.warning(f"Не удалось загрузить параметры сканирования: {e}")
        
        # Находим все файлы Beam№*.xlsx
        beam_files = []
        for filename in os.listdir(save_dir):
            if filename.startswith('Beam№') and filename.endswith('.xlsx'):
                try:
                    beam_num = int(filename.replace('Beam№', '').replace('.xlsx', ''))
                    beam_files.append((beam_num, os.path.join(save_dir, filename)))
                except ValueError:
                    continue
        
        if not beam_files:
            logger.warning(f"Не найдено файлов лучей в {save_dir}")
            return None
        
        beam_files.sort(key=lambda x: x[0])  # Сортируем по номеру луча
        
        # Используем параметры из JSON, если есть, иначе определяем из файлов
        if loaded_params:
            beams = loaded_params.get('beams', [beam_num for beam_num, _ in beam_files])
            freq_list = loaded_params.get('freq_list', [])
            x_list = loaded_params.get('x_list', [])
            y_list = loaded_params.get('y_list', [])
            step_x = loaded_params.get('step_x', 1.0)
            step_y = loaded_params.get('step_y', 1.0)
        else:
            beams = [beam_num for beam_num, _ in beam_files]
            freq_list = []
            x_list = []
            y_list = []
            step_x = 1.0
            step_y = 1.0
        
        # Загружаем данные из первого файла для определения структуры
        n_files = len(beam_files)
        first_beam_num, first_file = beam_files[0]
        root.emit(0.0, f'Открытие файла 1 из {n_files}…')
        workbook = load_workbook(first_file)
        sheet = workbook.active
        root.emit(0.0, 'Разбор структуры файла…')

        # Если частоты не загружены из JSON, определяем из файла
        if not freq_list:
            # Ищем все частоты
            row = 1
            while row <= sheet.max_row:
                cell = sheet.cell(row, 1)
                if cell.value == 'Frequency':
                    freq_cell = sheet.cell(row, 2)
                    if freq_cell.value:
                        try:
                            freq = float(freq_cell.value)
                            freq_list.append(freq)
                        except (ValueError, TypeError):
                            pass
                row += 1
            
            if not freq_list:
                logger.error("Не найдено частот в файле")
                return None
        
        # Определяем размеры данных из первой частоты
        first_freq_row = None
        for row in range(1, sheet.max_row + 1):
            if sheet.cell(row, 1).value == 'Frequency' and sheet.cell(row, 2).value == freq_list[0]:
                first_freq_row = row
                break
        
        if first_freq_row is None:
            logger.error("Не найдена первая частота в файле")
            return None
        
        # Находим размер данных (количество столбцов с данными)
        magnitude_row = first_freq_row + 1
        max_col = 0
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(magnitude_row + 1, col)  # Первая строка данных
            if cell.value is not None:
                max_col = max(max_col, col)
        
        len_y = max_col  # Количество столбцов = количество Y координат
        
        # Находим количество строк данных (до следующего Frequency или до конца)
        next_freq_row = None
        for row in range(first_freq_row + 1, sheet.max_row + 1):
            if sheet.cell(row, 1).value == 'Frequency':
                next_freq_row = row
                break
        
        if next_freq_row:
            # size_freq_data = 3 + len_x * 2
            # next_freq_row = first_freq_row + size_freq_data
            # len_x = (next_freq_row - first_freq_row - 3) // 2
            len_x = (next_freq_row - first_freq_row - 3) // 2
        else:
            # Последняя частота - считаем до конца
            phase_start_row = None
            for row in range(first_freq_row + 1, sheet.max_row + 1):
                if sheet.cell(row, 1).value == 'Phase':
                    phase_start_row = row
                    break
            if phase_start_row:
                # Заголовок 'Phase' идёт сразу за блоком амплитуды:
                # phase_start_row = first_freq_row + 2 + len_x
                len_x = phase_start_row - first_freq_row - 2
            else:
                # Если не нашли Phase, считаем по последней строке листа:
                # max_row = first_freq_row + 2 + len_x * 2
                len_x = (sheet.max_row - first_freq_row - 2) // 2
        
        size_freq_data = 3 + len_x * 2
        
        # Если координаты не загружены из JSON, используем индексы
        if not x_list:
            x_list = list(range(len_x))
        if not y_list:
            y_list = list(range(len_y))
        
        # Загружаем данные для всех лучей
        data = {}
        for idx, (beam_num, file_path) in enumerate(beam_files):
            part = root.part(idx / n_files, (idx + 1) / n_files)
            label = f'Луч {beam_num} ({idx + 1}/{n_files})'
            if idx == 0:
                beam_sheet = sheet  # первый файл уже открыт при разборе структуры
            else:
                part.emit(0.0, f'{label} — открытие файла…')
                beam_sheet = load_workbook(file_path).active
            data[beam_num] = _read_sheet_data(
                beam_sheet, freq_list, len_x, len_y, size_freq_data, x_list, y_list,
                progress=part, label=label)
        root.emit(1.0, 'Загрузка завершена')

        logger.info(f"Загружены данные: {len(beams)} лучей, {len(freq_list)} частот, {len_x}x{len_y} точек")
        
        result = {
            'beams': beams,
            'freq_list': freq_list,
            'data': data,
            'x_list': x_list,
            'y_list': y_list,
            'step_x': step_x,
            'step_y': step_y,
            'save_dir': save_dir
        }
        
        # Добавляем параметры из JSON, если они были загружены
        if loaded_params:
            result['left_x'] = loaded_params.get('left_x')
            result['right_x'] = loaded_params.get('right_x')
            result['up_y'] = loaded_params.get('up_y')
            result['down_y'] = loaded_params.get('down_y')
            # Добавляем настройки PNA и синхронизатора
            result['pna_settings'] = loaded_params.get('pna_settings')
            result['sync_settings'] = loaded_params.get('sync_settings')
        
        return result

    except LoadCancelled:
        logger.info(f"Загрузка папки отменена пользователем: {save_dir}")
        raise
    except Exception as e:
        logger.error(f"Ошибка при загрузке результатов измерения лучей: {e}", exc_info=True)
        return None


def _detect_freqs(sheet) -> list:
    """Список частот из листа (ячейки «Frequency» в первом столбце)."""
    freqs = []
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row, 1).value == 'Frequency':
            v = sheet.cell(row, 2).value
            try:
                freqs.append(float(v))
            except (TypeError, ValueError):
                pass
    return freqs


def _detect_layout(sheet, freq_list) -> Optional[tuple]:
    """Размеры данных листа: (len_x, len_y, size_freq_data).

    Логика та же, что и в load_beam_pattern_results: один блок частоты —
    это 3 строки заголовков + len_x строк амплитуды + len_x строк фазы.
    """
    first_freq_row = None
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row, 1).value == 'Frequency' and sheet.cell(row, 2).value == freq_list[0]:
            first_freq_row = row
            break
    if first_freq_row is None:
        logger.error("Не найдена первая частота в файле")
        return None

    magnitude_row = first_freq_row + 1
    len_y = 0
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(magnitude_row + 1, col).value is not None:
            len_y = max(len_y, col)

    next_freq_row = None
    for row in range(first_freq_row + 1, sheet.max_row + 1):
        if sheet.cell(row, 1).value == 'Frequency':
            next_freq_row = row
            break

    if next_freq_row:
        len_x = (next_freq_row - first_freq_row - 3) // 2
    else:
        phase_start_row = None
        for row in range(first_freq_row + 1, sheet.max_row + 1):
            if sheet.cell(row, 1).value == 'Phase':
                phase_start_row = row
                break
        if phase_start_row:
            # Заголовок 'Phase' идёт сразу за блоком амплитуды:
            # phase_start_row = first_freq_row + 2 + len_x
            len_x = phase_start_row - first_freq_row - 2
        else:
            # max_row = first_freq_row + 2 + len_x * 2
            len_x = (sheet.max_row - first_freq_row - 2) // 2

    size_freq_data = 3 + len_x * 2
    return len_x, len_y, size_freq_data


def _read_sheet_data(sheet, freq_list, len_x, len_y, size_freq_data, x_list, y_list,
                     progress=None, label='') -> dict:
    """Прочитать амплитуду и фазу по всем частотам с одного листа.

    progress — необязательный отчёт о ходе чтения (доля 0..1 внутри листа);
    он же проверяет отмену, поэтому длинное чтение можно прервать.

    Returns: {freq: {'x', 'y', 'amp', 'phase'}}.
    """
    data = {}
    n_freq = max(len(freq_list), 1)
    n_rows = max(len_x, 1)

    def tick(freq, freq_idx, half, x_idx):
        """Доля: частоты × две половины блока (half=0 — амплитуда, 1 — фаза)."""
        if progress is None or x_idx % 8:
            return
        frac = (freq_idx + (half + x_idx / n_rows) / 2) / n_freq
        text = f'{freq:g} МГц' if not label else f'{label} — {freq:g} МГц'
        progress.emit(frac, text)

    for freq_idx, freq in enumerate(freq_list):
        row_start = freq_idx * size_freq_data + 1
        amp_2d = np.full((len_y, len_x), np.nan)
        phase_2d = np.full((len_y, len_x), np.nan)

        for x_idx in range(len_x):
            tick(freq, freq_idx, 0, x_idx)
            for y_idx in range(len_y):
                cell = sheet.cell(row_start + 2 + x_idx, y_idx + 1)
                if cell.value is not None:
                    try:
                        amp_2d[y_idx, x_idx] = float(cell.value)
                    except (ValueError, TypeError):
                        pass

        for x_idx in range(len_x):
            tick(freq, freq_idx, 1, x_idx)
            for y_idx in range(len_y):
                cell = sheet.cell(row_start + 3 + len_x + x_idx, y_idx + 1)
                if cell.value is not None:
                    try:
                        phase_2d[y_idx, x_idx] = float(cell.value)
                    except (ValueError, TypeError):
                        pass

        data[freq] = {
            'x': x_list,
            'y': y_list,
            'amp': amp_2d.tolist(),
            'phase': phase_2d.tolist(),
        }
    return data


class BeamFileFormatError(Exception):
    """Файл не соответствует ожидаемому формату измерения Beam№*.xlsx.

    Сообщение исключения предназначено для показа пользователю.
    """


def load_single_beam_file(file_path: str, on_progress=None,
                          is_cancelled=None) -> dict:
    """Загрузить ОДИН файл измерения (Beam№*.xlsx) для отдельного пересчёта.

    В отличие от load_beam_pattern_results, не требует папки и scan_params.json:
    частоты берутся из самого файла, шаг сканера (dx/dy) и прочие параметры
    задаются пользователем вручную. «Луч» здесь — имя файла.

    Принять можно ЛЮБОЙ файл, но если внутри не тот формат — поднимается
    BeamFileFormatError с понятным сообщением (что именно не так).

    on_progress(frac, text) — ход загрузки (frac ∈ 0..1 или None, если доля
    неизвестна); is_cancelled() -> bool — пользователь нажал «Отмена».

    Returns: {
        'name': имя файла без расширения (роль «луча»),
        'freq_list': [...],
        'data': {name: {freq: {'x','y','amp','phase'}}},
        'x_list', 'y_list', 'file_path'
    }

    Raises:
        BeamFileFormatError — файл не открывается как .xlsx или не содержит
        ожидаемой структуры измерения.
        LoadCancelled — загрузка прервана пользователем.
    """
    if not os.path.isfile(file_path):
        raise BeamFileFormatError(f'Файл не найден:\n{file_path}')

    progress = _Progress(on_progress, is_cancelled)
    # Разбор книги идёт внутри openpyxl одним куском — доля неизвестна, поэтому
    # полоса на это время «бегущая», а не процентная.
    progress.busy(f'Открытие файла {os.path.basename(file_path)}…')

    try:
        workbook = load_workbook(file_path)
    except Exception as e:
        logger.warning(f"Не удалось открыть как Excel: {file_path}: {e}")
        raise BeamFileFormatError(
            'Не удалось открыть файл как таблицу Excel (.xlsx).\n'
            'Имя файла может быть любым, но внутри должен быть формат '
            'измерения (как у файлов из режима «Измерение лучей АФАР»).') from e

    sheet = workbook.active
    if sheet is None:
        raise BeamFileFormatError('В книге Excel нет активного листа с данными.')

    progress.busy('Разбор структуры файла…')
    freq_list = _detect_freqs(sheet)
    if not freq_list:
        raise BeamFileFormatError(
            'Файл не похож на измерение: не найдены строки «Frequency».\n'
            'Имя файла роли не играет — важен формат данных внутри '
            '(как в режиме «Измерение лучей АФАР»).')

    layout = _detect_layout(sheet, freq_list)
    if layout is None:
        raise BeamFileFormatError(
            'Не удалось определить структуру данных: в файле есть «Frequency», '
            'но не найден блок амплитуды/фазы ожидаемого вида.')
    len_x, len_y, size_freq_data = layout
    if len_x <= 0 or len_y <= 0:
        raise BeamFileFormatError(
            'Не удалось определить размер сетки данных (амплитуда/фаза).')

    x_list = list(range(len_x))
    y_list = list(range(len_y))
    data_by_freq = _read_sheet_data(
        sheet, freq_list, len_x, len_y, size_freq_data, x_list, y_list,
        progress=progress, label='Чтение данных')
    progress.emit(1.0, 'Загрузка завершена')

    # Должны быть хоть какие-то числовые значения амплитуды.
    has_amp = any(
        np.any(np.isfinite(np.asarray(fd['amp'], dtype=float)))
        for fd in data_by_freq.values()
    )
    if not has_amp:
        raise BeamFileFormatError(
            'Файл распознан как измерение, но не содержит числовых данных '
            'амплитуды.')

    name = os.path.splitext(os.path.basename(file_path))[0]
    logger.info(f"Загружен файл: {name}, {len(freq_list)} частот, {len_x}x{len_y} точек")
    return {
        'name': name,
        'freq_list': freq_list,
        'data': {name: data_by_freq},
        'x_list': x_list,
        'y_list': y_list,
        'file_path': file_path,
    }
