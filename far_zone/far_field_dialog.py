# -*- coding: utf-8 -*-
"""Утилита расчёта диаграммы направленности в дальней зоне (NF -> FF).

Вход — папка с результатами режима «Измерение лучей АФАР» (Beam№*.xlsx +
scan_params.json) ЛИБО отдельный файл измерения (кнопка «Открыть файл»).
В режиме одиночного файла роль «луча» играет имя файла, а dx/dy задаются вручную.

Два вида, переключаются кнопками в верхней панели:
- «Ближнее поле» — 2D-карта измеренного поля (амплитуда или фаза по апертуре);
  доступна сразу после загрузки, см. near_field_panel;
- «Главные сечения ДН» — дальняя зона; включается только после пересчёта
  (кнопка «Пересчитать…» -> окно параметров -> фоновый счёт с кэшем).
Луч и частота переключаются общими комбобоксами: в ближнем поле доступно всё
загруженное, в дальней зоне — только рассчитанное.

В дальней зоне амплитуда и фаза показываются НА ОДНОМ графике: 4 трассы (Az/El ×
амплитуда/фаза), две оси Y — слева амплитуда (дБ), справа фаза (°). Любую трассу
можно скрыть тумблером. Поддержка: значение при наведении, перетаскиваемые линии-
маркеры V/H со значениями на пересечении, наложение трасс, нормировка (только
амплитуда), поиск максимумов (только амплитуда), экспорт PNG/CSV.

Единицы шага сканера — сантиметры (как в MATLAB): dx/dy [см] -> метры (×1e-2).
"""

import os
import csv

import numpy as np
import pyqtgraph as pg
import pyqtgraph.exporters  # noqa: F401  (регистрирует pg.exporters)
from PyQt5 import QtCore, QtGui, QtWidgets
from loguru import logger

from .nf2ff_solver import solve_sections, find_peak_indices, side_lobe_level
from .beam_loader import (load_beam_pattern_results, load_single_beam_file,
                          BeamFileFormatError, LoadCancelled)
from .beam_mapping import (beam_to_angles, order_beams, BEAM_ORDER_NUMBER,
                           BEAM_ORDER_AZIMUTH, BEAM_ORDER_ELEVATION)
from .near_field_panel import NearFieldPanel
from .app_style import reserve_bold_width
from .icon_utils import set_button_icon, app_icon
from .design_tokens import (ICON_DEFAULT, ICON_MD, ICON_ON_ACCENT, ICON_SM,
                            MARKER_H, MARKER_PEAK, MARKER_V, MASK_LINE,
                            OVERLAY_COLORS, PLOT_AXIS_MUTED, PLOT_BG,
                            PLOT_TITLE, STATUS_ICON, TRACE_AZ_AMP,
                            TRACE_AZ_PHASE, TRACE_EL_AMP, TRACE_EL_PHASE)

# Цвета берутся из design_tokens (там же видно, почему они разведены именно так).
AZ_COLOR = TRACE_AZ_AMP
EL_COLOR = TRACE_EL_AMP
PHASE_AZ_COLOR = TRACE_AZ_PHASE
PHASE_EL_COLOR = TRACE_EL_PHASE
MASK_COLOR = MASK_LINE                 # линия маски УБЛ
MARKER_V_COLOR = MARKER_V              # обычный вертикальный маркер
MARKER_H_COLOR = MARKER_H              # обычный горизонтальный маркер
PEAK_COLOR = MARKER_PEAK               # маркер поиска максимума (магнитится к пикам)
MARKER_WIDTH = 1.8                     # толщина линии маркеров (была 1)
PEAK_WIDTH = 2.4                       # маркер-максимум чуть толще обычных
TRACE_WIDTH = 2.4                      # амплитуда — сплошная
PHASE_WIDTH = 2.0                      # фаза — штриховая, чуть тоньше
DEG = np.pi / 180
STEP_TO_M = 1e-2  # шаг сканера: см -> метры


def angle_axis(left, right, step):
    """Сетка углов (град) — той же формулой, что и солвер (совпадает по длине)."""
    return np.degrees(np.arange(left * DEG, right * DEG + step * DEG, step * DEG))


def _csv_num(value):
    """Число для CSV; углы вне действительного окна (NaN) — пустая ячейка.

    Писать «nan» текстом нельзя: Excel принимает такую колонку за строковую и
    не строит по ней график. Пустая ячейка — как в экспорте карты ближнего поля.
    """
    return '' if not np.isfinite(value) else f'{value:.4f}'


# ============================================================ Окно параметров
class FarFieldParamsDialog(QtWidgets.QDialog):
    """Параметры пересчёта + выбор лучей и частот."""

    def __init__(self, beams, freqs, defaults=None, single_file=False, parent=None):
        super().__init__(parent)
        self.setWindowTitle('Параметры пересчёта')
        self.setMinimumWidth(560)
        defaults = defaults or {}
        self._single_file = single_file

        root = QtWidgets.QVBoxLayout(self)
        root.setSpacing(10)

        params_group = QtWidgets.QGroupBox('Параметры расчёта')
        grid = QtWidgets.QGridLayout(params_group)
        self.az_from = self._dspin(-90, 90, defaults.get('az_from', -5), ' °')
        self.az_to = self._dspin(-90, 90, defaults.get('az_to', 5), ' °')
        self.az_step = self._dspin(0.001, 5, defaults.get('az_step', 0.05), ' °', 3)
        self.el_from = self._dspin(-90, 90, defaults.get('el_from', -20), ' °')
        self.el_to = self._dspin(-90, 90, defaults.get('el_to', 20), ' °')
        self.el_step = self._dspin(0.001, 5, defaults.get('el_step', 0.05), ' °', 3)
        self.dx_spin = self._dspin(0.01, 1000, defaults.get('dx', 1.0), ' см', 3)
        self.dy_spin = self._dspin(0.01, 1000, defaults.get('dy', 1.0), ' см', 3)

        self.fft_combo = QtWidgets.QComboBox()
        for text, val in (('×2 (быстро)', 2), ('×4', 4), ('×8', 8), ('×16 (точно)', 16)):
            self.fft_combo.addItem(text, val)
        self.fft_combo.setToolTip(
            'Дополнение нулями перед БПФ (число точек = ближайшая степень 2 × множитель).\n'
            'Больше — плотнее сетка и точнее интерполяция, но дольше и больше памяти.')
        pad_idx = self.fft_combo.findData(int(defaults.get('pad', 4)))
        self.fft_combo.setCurrentIndex(pad_idx if pad_idx >= 0 else 1)

        self.interp_combo = QtWidgets.QComboBox()
        self.interp_combo.addItem('Бикубическая', 'cubic')
        self.interp_combo.addItem('Кубический сплайн', 'spline')
        self.interp_combo.addItem('Билинейная', 'linear')
        self.interp_combo.setToolTip(
            'Метод интерполяции спектра на углы.\n'
            '• Кубический сплайн — как interp2(...,\'spline\') в MATLAB (точнее всего);\n'
            '• Бикубическая — кубическая свёртка (interp2 \'cubic\'), близко к сплайну;\n'
            '• Билинейная — быстрее/грубее.')
        i_idx = self.interp_combo.findData(str(defaults.get('interp', 'cubic')))
        self.interp_combo.setCurrentIndex(i_idx if i_idx >= 0 else 0)

        self.abs_combo = QtWidgets.QComboBox()
        self.abs_combo.addItem('После (комплекс → |·|)', 'after')
        self.abs_combo.addItem('До (модуль |fx| → интерп.)', 'before')
        self.abs_combo.setToolTip(
            'Когда брать модуль амплитуды:\n'
            '• После — интерполируется комплексный спектр, затем модуль '
            '(точнее у нулей, наш метод);\n'
            '• До — модуль берётся до интерполяции (как в MATLAB).\n'
            'Фаза всегда считается по комплексу.')
        a_idx = self.abs_combo.findData(str(defaults.get('abs_mode', 'after')))
        self.abs_combo.setCurrentIndex(a_idx if a_idx >= 0 else 0)

        grid.addWidget(QtWidgets.QLabel('Az:'), 0, 0)
        grid.addWidget(self._labeled('от', self.az_from), 0, 1)
        grid.addWidget(self._labeled('до', self.az_to), 0, 2)
        grid.addWidget(self._labeled('шаг', self.az_step), 0, 3)
        grid.addWidget(QtWidgets.QLabel('El:'), 1, 0)
        grid.addWidget(self._labeled('от', self.el_from), 1, 1)
        grid.addWidget(self._labeled('до', self.el_to), 1, 2)
        grid.addWidget(self._labeled('шаг', self.el_step), 1, 3)
        grid.addWidget(self._labeled('Шаг X (dx)', self.dx_spin), 2, 1)
        grid.addWidget(self._labeled('Шаг Y (dy)', self.dy_spin), 2, 2)
        grid.addWidget(self._labeled('Точек БПФ', self.fft_combo), 2, 3)
        grid.addWidget(self._labeled('Интерполяция', self.interp_combo), 3, 1)
        grid.addWidget(self._labeled('Модуль (abs)', self.abs_combo), 3, 2, 1, 2)
        root.addWidget(params_group)

        lists_row = QtWidgets.QHBoxLayout()
        if single_file:
            beams_group, self.beams_list = self._build_check_list('Файл', beams, lambda b: str(b))
        else:
            beams_group, self.beams_list = self._build_check_list('Лучи', beams, lambda b: f'Луч {b}')
        freqs_group, self.freqs_list = self._build_check_list('Частоты', freqs, lambda f: f'{f:g} МГц')
        lists_row.addWidget(beams_group)
        lists_row.addWidget(freqs_group)
        root.addLayout(lists_row)

        buttons = QtWidgets.QDialogButtonBox(
            QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel)
        ok_button = buttons.button(QtWidgets.QDialogButtonBox.Ok)
        ok_button.setText('Рассчитать')
        # Главное действие окна — заливка акцентом (см. theme.qss, primary).
        ok_button.setProperty('primary', True)
        ok_button.setDefault(True)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        root.addWidget(buttons)

    @staticmethod
    def _dspin(lo, hi, val, suffix='', decimals=2):
        spin = QtWidgets.QDoubleSpinBox()
        spin.setRange(lo, hi)
        spin.setDecimals(decimals)
        spin.setValue(val)
        if suffix:
            spin.setSuffix(suffix)
        return spin

    @staticmethod
    def _labeled(text, widget):
        box = QtWidgets.QWidget()
        lay = QtWidgets.QVBoxLayout(box)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(2)
        lbl = QtWidgets.QLabel(text)
        lbl.setStyleSheet('color: #667085; font-size: 8pt;')
        lay.addWidget(lbl)
        lay.addWidget(widget)
        return box

    def _build_check_list(self, title, items, fmt):
        group = QtWidgets.QGroupBox(title)
        layout = QtWidgets.QVBoxLayout(group)
        lst = QtWidgets.QListWidget()
        lst.setSelectionMode(QtWidgets.QAbstractItemView.NoSelection)
        for value in items:
            item = QtWidgets.QListWidgetItem(fmt(value))
            item.setFlags(item.flags() | QtCore.Qt.ItemIsUserCheckable)
            item.setCheckState(QtCore.Qt.Checked)
            item.setData(QtCore.Qt.UserRole, value)
            lst.addItem(item)
        layout.addWidget(lst)

        btn_row = QtWidgets.QHBoxLayout()
        for text, action in (('Все', QtCore.Qt.Checked), ('Снять', QtCore.Qt.Unchecked), ('Инверт.', None)):
            btn = QtWidgets.QPushButton(text)
            if action is None:
                btn.clicked.connect(lambda _, w=lst: self._invert(w))
            else:
                btn.clicked.connect(lambda _, w=lst, s=action: self._set_all(w, s))
            btn_row.addWidget(btn)
        layout.addLayout(btn_row)
        return group, lst

    @staticmethod
    def _set_all(lst, state):
        for i in range(lst.count()):
            lst.item(i).setCheckState(state)

    @staticmethod
    def _invert(lst):
        for i in range(lst.count()):
            item = lst.item(i)
            item.setCheckState(QtCore.Qt.Unchecked if item.checkState() == QtCore.Qt.Checked else QtCore.Qt.Checked)

    @staticmethod
    def _checked_values(lst):
        return [lst.item(i).data(QtCore.Qt.UserRole)
                for i in range(lst.count()) if lst.item(i).checkState() == QtCore.Qt.Checked]

    def selected_beams(self):
        return self._checked_values(self.beams_list)

    def selected_freqs(self):
        return self._checked_values(self.freqs_list)

    def params(self):
        return {
            'az_from': self.az_from.value(), 'az_to': self.az_to.value(), 'az_step': self.az_step.value(),
            'el_from': self.el_from.value(), 'el_to': self.el_to.value(), 'el_step': self.el_step.value(),
            'dx': self.dx_spin.value(), 'dy': self.dy_spin.value(),
            'pad': int(self.fft_combo.currentData()),
            'interp': str(self.interp_combo.currentData()),
            'abs_mode': str(self.abs_combo.currentData()),
        }

    def accept(self):
        if self.az_from.value() >= self.az_to.value() or self.el_from.value() >= self.el_to.value():
            QtWidgets.QMessageBox.warning(self, 'Проверьте диапазоны', '«от» должно быть меньше «до» по Az и El.')
            return
        if not self.selected_beams():
            if self._single_file:
                QtWidgets.QMessageBox.warning(self, 'Нет файла', 'Отметьте файл для пересчёта.')
            else:
                QtWidgets.QMessageBox.warning(self, 'Нет лучей', 'Выберите хотя бы один луч.')
            return
        if not self.selected_freqs():
            QtWidgets.QMessageBox.warning(self, 'Нет частот', 'Выберите хотя бы одну частоту.')
            return
        super().accept()


# ============================================================ Фоновая загрузка
class _LoadWorker(QtCore.QObject):
    """Чтение Excel в отдельном потоке: окно не «замирает», виден прогресс.

    Разбор книги openpyxl занимает секунды-минуты, поэтому загрузка вынесена из
    UI-потока, а загрузчику передаются callback'и прогресса и отмены.
    """

    progress = QtCore.pyqtSignal(int, str)   # промилле (0..1000), -1 — доля неизвестна
    finished = QtCore.pyqtSignal(object)     # результат загрузчика
    failed = QtCore.pyqtSignal(str, str)     # заголовок окна, текст
    cancelled = QtCore.pyqtSignal()
    done = QtCore.pyqtSignal()

    def __init__(self, kind, path):
        super().__init__()
        self.kind = kind            # 'dir' — папка лучей, 'file' — один файл
        self.path = path
        self._stop = False
        self._last = None           # последнее отправленное (промилле, текст)

    def stop(self):
        self._stop = True

    def _on_progress(self, frac, text):
        """Отсеять повторы: тиков много, а полосе нужны только изменения."""
        permille = -1 if frac is None else int(round(frac * 1000))
        if (permille, text) == self._last:
            return
        self._last = (permille, text)
        self.progress.emit(permille, text)

    @QtCore.pyqtSlot()
    def run(self):
        try:
            if self.kind == 'dir':
                result = load_beam_pattern_results(
                    self.path, on_progress=self._on_progress,
                    is_cancelled=lambda: self._stop)
                if not result or not result.get('data'):
                    self.failed.emit('Ошибка',
                                     'Не удалось загрузить данные из выбранной папки.')
                    return
            else:
                result = load_single_beam_file(
                    self.path, on_progress=self._on_progress,
                    is_cancelled=lambda: self._stop)
            self.finished.emit(result)
        except LoadCancelled:
            self.cancelled.emit()
        except BeamFileFormatError as exc:
            self.failed.emit('Неподходящий файл',
                             f'{os.path.basename(self.path)}\n\n{exc}')
        except Exception as exc:
            logger.error(f'Ошибка чтения {self.path}: {exc}', exc_info=True)
            self.failed.emit('Ошибка', f'Не удалось прочитать данные:\n{exc}')
        finally:
            self.done.emit()


# ================================================================ Фоновый счёт
class _FarFieldWorker(QtCore.QObject):
    progress = QtCore.pyqtSignal(int, int, str)
    finished = QtCore.pyqtSignal(dict)
    failed = QtCore.pyqtSignal(str)
    done = QtCore.pyqtSignal()

    def __init__(self, data, tasks, params):
        super().__init__()
        self.data = data
        self.tasks = tasks
        self.p = params
        self._stop = False

    def stop(self):
        self._stop = True

    @QtCore.pyqtSlot()
    def run(self):
        cache = {}
        total = len(self.tasks)
        try:
            for i, (beam, freq) in enumerate(self.tasks, 1):
                if self._stop:
                    break
                entry = self._compute_one(beam, freq)
                if entry is not None:
                    cache[(beam, freq)] = entry
                self.progress.emit(i, total, f'Луч {beam}, {freq:g} МГц')
            self.finished.emit(cache)
        except Exception as exc:
            logger.error(f'Ошибка фонового расчёта дальней зоны: {exc}', exc_info=True)
            self.failed.emit(str(exc))
        finally:
            self.done.emit()

    def _compute_one(self, beam, freq):
        fd = self.data.get(beam, {}).get(freq)
        if not fd:
            return None
        amp = np.asarray(fd['amp'], dtype=float).T      # (y,x) -> (x,y)
        phase = np.asarray(fd['phase'], dtype=float).T
        if np.all(np.isnan(amp)):
            return None
        p = self.p
        return solve_sections(
            amp, phase, freq * 1e6, p['dx'] * STEP_TO_M, p['dy'] * STEP_TO_M,
            p['az_from'], p['az_to'], p['az_step'],
            p['el_from'], p['el_to'], p['el_step'],
            p.get('pad', 4), p.get('interp', 'cubic'), p.get('abs_mode', 'after'),
        )


# ============================================================ Панель графика
class FarFieldPlotPanel(QtWidgets.QWidget):
    """График с двумя осями Y: слева амплитуда (дБ), справа фаза (°).

    trace_defs: список (name, color, axis, dashed), axis ∈ {'L','R'} —
    левая/правая ось. Левые трассы считаются амплитудными (к ним применяются
    нормировка и поиск максимумов). Линии-маркеры (верт./гориз.) показывают
    значения пересечения с трассами: ПКМ по линии — удалить, двойной клик —
    ввести точное значение.

    Кодировка: ТОН = плоскость (Az/El), НАЧЕРТАНИЕ = величина (амплитуда
    сплошная, фаза штриховая). Раньше все четыре трассы были сплошными и
    различались только цветом — по какой оси читать кривую, приходилось
    вспоминать.
    """

    def __init__(self, trace_defs, left_unit='дБ', left_range=(-60, 0),
                 right_unit='°', right_range=(-200, 200), parent=None):
        super().__init__(parent)
        self._trace_defs = list(trace_defs)
        self._left_unit = left_unit
        self._right_unit = right_unit
        self._left_range = left_range
        self._right_range = right_range

        self._label_font = QtGui.QFont()
        self._label_font.setPointSize(10)
        self._label_font.setBold(True)

        self._normalize = True
        self._cursors = []      # [{line, kind, dotsL, dotsR, labelsL, labelsR}]
        self._overlays = []     # [[curve, vb, name]]
        self._hover_text = ''
        self._overlay_count = 0
        self._vis_btns = {}
        self._peak_cursor = None
        self._active_cursor = None   # выделенная строка таблицы = подсвеченная линия
        self._mask_line = None
        self._mask_db = None

        root = QtWidgets.QHBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(8)
        root.addWidget(self._build_tools(), 0)

        plot_area = QtWidgets.QVBoxLayout()
        plot_area.setContentsMargins(0, 0, 0, 0)
        plot_area.setSpacing(2)
        plot_area.addLayout(self._build_plane_bar())

        self._build_plot(left_unit, left_range, right_unit, right_range)
        plot_area.addWidget(self.plot, 1)

        self.readout = QtWidgets.QLabel(' ')
        self.readout.setObjectName('ffReadout')
        self.readout.setMinimumHeight(16)
        plot_area.addWidget(self.readout)
        plot_area.addWidget(self._build_marker_table())
        root.addLayout(plot_area, 1)

        self._build_traces(trace_defs)
        self._build_hover()

        self._proxy = pg.SignalProxy(self.plot.scene().sigMouseMoved, rateLimit=60, slot=self._on_mouse_move)
        self.plot.scene().sigMouseClicked.connect(self._on_scene_click)

    # ------------------------------------------------------------- построение
    def _build_plot(self, left_unit, left_range, right_unit, right_range):
        self.plot = pg.PlotWidget()
        self.plot.setBackground(PLOT_BG)
        pi = self.plot.getPlotItem()
        self._pi = pi
        pi.setTitle('Дальняя зона: амплитуда (дБ) · фаза (°)', color=PLOT_TITLE, size='11pt')
        pi.showGrid(x=True, y=True, alpha=0.22)
        pi.setMenuEnabled(False)
        pi.setLabel('bottom', 'Угол, °')
        pi.setLabel('left', left_unit)
        self.vbL = pi.vb
        self.vbL.setYRange(*left_range)

        # Вторая ось Y (фаза) — отдельный ViewBox, связанный по X.
        self.vbR = pg.ViewBox()
        pi.showAxis('right')
        pi.scene().addItem(self.vbR)
        pi.getAxis('right').linkToView(self.vbR)
        self.vbR.setXLink(pi)
        self.vbR.setMouseEnabled(x=False, y=False)
        # Ось нейтральная: на ней ДВЕ трассы (Az и El), и красить её в цвет
        # одной из них — значит утверждать неверное соответствие.
        # Принадлежность к оси теперь видна по штриховке трасс.
        ax_r = pi.getAxis('right')
        ax_r.setLabel(right_unit, color=PLOT_AXIS_MUTED)
        ax_r.setTextPen(pg.mkPen(PLOT_AXIS_MUTED))
        self.vbR.setYRange(*right_range)

        self.vbL.sigResized.connect(self._sync_views)
        QtCore.QTimer.singleShot(0, self._sync_views)

        self.legend = pi.addLegend(offset=(-12, 8))

    def _sync_views(self):
        self.vbR.setGeometry(self.vbL.sceneBoundingRect())
        self.vbR.linkedViewChanged(self.vbL, self.vbR.XAxis)

    def _build_traces(self, trace_defs):
        self._traces = []
        for name, color, axis, dashed in trace_defs:
            if dashed:
                pen = pg.mkPen(color, width=PHASE_WIDTH, style=QtCore.Qt.DashLine)
            else:
                pen = pg.mkPen(color, width=TRACE_WIDTH)
            # connect='finite' — рвать линию на NaN (обрезанные нефизичные участки ДН)
            curve = pg.PlotDataItem([], [], pen=pen, name=name, connect='finite')
            vb = self.vbL if axis == 'L' else self.vbR
            vb.addItem(curve)
            self.legend.addItem(curve, name)
            self._traces.append({
                'name': name, 'color': color, 'axis': axis, 'amp': axis == 'L',
                'unit': self._left_unit if axis == 'L' else self._right_unit,
                'curve': curve, 'vb': vb, 'visible': True,
                'x': np.array([]), 'y_raw': np.array([]), 'max_val': 0.0,
            })

    def _build_hover(self):
        self._hover = {}
        for key, vb in (('L', self.vbL), ('R', self.vbR)):
            dot = pg.ScatterPlotItem(size=11, brush=pg.mkBrush('#111827'), pen=pg.mkPen('w', width=1.5))
            dot.setZValue(50)
            vb.addItem(dot, ignoreBounds=True)
            lab = pg.TextItem(color='#111827', anchor=(0, 1))
            lab.setZValue(51)
            vb.addItem(lab, ignoreBounds=True)
            lab.hide()
            self._hover[key] = {'dot': dot, 'lab': lab}
        self.vline = pg.InfiniteLine(angle=90, movable=False,
                                     pen=pg.mkPen('#c7d0dc', style=QtCore.Qt.DashLine))
        self.vline.hide()
        self.plot.addItem(self.vline, ignoreBounds=True)

    # --------------------------------------------------- инструменты (слева)
    def _icon_btn(self, icon, tip, slot, checkable=False, shortcut=None,
                  color=ICON_DEFAULT, color_on=None):
        btn = QtWidgets.QToolButton()
        btn.setProperty('plotTool', True)
        btn.setIcon(app_icon(icon, color=color, color_on=color_on))
        btn.setIconSize(QtCore.QSize(ICON_MD, ICON_MD))
        if shortcut:
            btn.setShortcut(QtGui.QKeySequence(shortcut))
            tip = f'{tip}  [{shortcut}]'
        btn.setToolTip(tip)
        btn.setCheckable(checkable)
        btn.setAutoRaise(True)
        btn.setFixedSize(32, 32)
        btn.clicked.connect(slot)
        return btn

    @staticmethod
    def _tool_separator():
        """Тонкая черта между группами инструментов (маркеры | вид | экспорт)."""
        line = QtWidgets.QWidget()
        line.setFixedHeight(1)
        line.setStyleSheet('background-color: #e2e8f0;')
        return line

    def _build_tools(self):
        """Компактная вертикальная панель иконок-инструментов.

        Иконка красится в цвет того, что кнопка создаёт на графике: маркеры
        видно по цвету линии, а не по памяти. Служебные (масштаб, экспорт)
        остаются нейтральными — так группы отличаются сами собой.
        """
        panel = QtWidgets.QWidget()
        panel.setFixedWidth(40)
        col = QtWidgets.QVBoxLayout(panel)
        col.setContentsMargins(2, 2, 2, 2)
        col.setSpacing(4)

        col.addWidget(self._icon_btn('cursor-h', 'Добавить горизонтальный маркер (по амплитуде)',
                                     self._add_hcursor, shortcut='H', color=MARKER_H_COLOR))
        col.addWidget(self._icon_btn('cursor-v', 'Добавить вертикальный маркер',
                                     self._add_vcursor, shortcut='V', color=MARKER_V_COLOR))
        col.addWidget(self._icon_btn('max-global', 'Маркер максимума (отдельный цвет): в главный максимум амплитуды.\n'
                                                   'Перетаскивание маркера притягивается к ближайшему максимуму',
                                     self._marker_to_max, shortcut='M', color=PEAK_COLOR))
        col.addWidget(self._icon_btn('max-local', 'Маркер максимума: следующий локальный максимум по кругу.\n'
                                                  'Перетаскивание маркера притягивается к ближайшему максимуму',
                                     self._next_local_max, shortcut='Shift+M', color=PEAK_COLOR))
        col.addWidget(self._icon_btn('clear-markers', 'Убрать все маркеры',
                                     self.clear_annotations, shortcut='Shift+C'))
        col.addWidget(self._tool_separator())
        col.addWidget(self._icon_btn('autoscale', 'Автомасштаб по данным', self.autoscale, shortcut='A'))
        self.norm_btn = self._icon_btn('normalize', 'Нормировка амплитуды к максимуму (дБ)',
                                       self._toggle_norm, checkable=True, shortcut='N',
                                       color_on=AZ_COLOR)
        self.norm_btn.setChecked(self._normalize)
        col.addWidget(self.norm_btn)
        col.addWidget(self._tool_separator())
        col.addWidget(self._icon_btn('csv', 'Экспорт данных в CSV', self._export_csv))
        col.addWidget(self._icon_btn('png', 'Экспорт графика в PNG', self._export_png))
        col.addStretch(1)
        return panel

    # ------------------------------------------------------- таблица маркеров
    def _marker_columns(self):
        """Колонки таблицы: [(вид, имя трассы, цвет, штриховая)].

        Δ стоит СВОЕЙ колонкой сразу за каждой амплитудной трассой. Раньше
        дельта была одна на всю строку, а горизонтальный маркер режет оба
        главных сечения: ширины у них разные (в Az апертура втрое длиннее —
        луч уже), и в единственную ячейку попадала ширина того сечения, что
        считалось последним. Вторая молча пропадала.
        """
        cols = [('title', None, None, False), ('pos', None, None, False)]
        for name, color, axis, dashed in self._trace_defs:
            cols.append(('trace', name, color, dashed))
            if axis == 'L':          # Δ есть только у амплитуд: уровень маркера — по левой оси
                cols.append(('delta', name, color, dashed))
        return cols

    @staticmethod
    def _column_title(kind, name):
        if kind == 'title':
            return 'Маркер'
        if kind == 'pos':
            return 'Позиция'
        if kind == 'trace':
            return name
        return 'Δ ' + name.split()[0]      # «Az ампл.» -> «Δ Az»

    @staticmethod
    def _line_swatch(color, dashed=False, width=22, height=12):
        """Чип с отрезком линии — как в легенде: тон и начертание.

        Цвет текста в ячейке съедается на светлых оттенках и не виден в
        заголовке (его красит QSS), а чип читается одинаково везде.
        """
        # Рисуем в физических пикселях экрана: на 125–150 % масштабе Windows
        # чип 22×12 иначе растягивается и мылится
        screen = QtWidgets.QApplication.primaryScreen() if QtWidgets.QApplication.instance() else None
        ratio = float(screen.devicePixelRatio()) if screen is not None else 1.0
        pm = QtGui.QPixmap(int(width * ratio), int(height * ratio))
        pm.setDevicePixelRatio(ratio)
        pm.fill(QtCore.Qt.transparent)
        painter = QtGui.QPainter(pm)
        painter.setRenderHint(QtGui.QPainter.Antialiasing)
        pen = QtGui.QPen(QtGui.QColor(color))
        pen.setWidthF(2.4)
        pen.setCapStyle(QtCore.Qt.FlatCap)
        if dashed:
            pen.setStyle(QtCore.Qt.CustomDashLine)
            pen.setDashPattern([2.0, 1.6])
        painter.setPen(pen)
        painter.drawLine(1, height // 2, width - 1, height // 2)
        painter.end()
        return QtGui.QIcon(pm)

    def _swatch(self, color, dashed=False):
        """Чип с кэшем: таблица перерисовывается на каждый сдвиг маркера."""
        key = (color, dashed)
        if key not in self._swatches:
            self._swatches[key] = self._line_swatch(color, dashed)
        return self._swatches[key]

    def _build_marker_table(self):
        """Значения маркеров таблицей, как в анализаторах цепей.

        Раньше всё сваливалось в одну строку под графиком: с четырьмя трассами
        и парой курсоров она была длиннее окна, а QLabel режет текст без
        многоточия — конец просто пропадал.
        """
        self._marker_cols = self._marker_columns()
        self._swatches = {}
        table = QtWidgets.QTableWidget(0, len(self._marker_cols))
        table.setObjectName('ffMarkerTable')
        table.setHorizontalHeaderLabels(
            [self._column_title(kind, name) for kind, name, _c, _d in self._marker_cols])
        for column, (kind, name, color, dashed) in enumerate(self._marker_cols):
            head = table.horizontalHeaderItem(column)
            if kind == 'trace':
                head.setIcon(self._swatch(color, dashed))
                head.setToolTip(f'«{name}» в точке маркера.\n'
                                'Для горизонтального маркера — углы пересечения трассы с его уровнем')
            elif kind == 'delta':
                head.setIcon(self._swatch(color, dashed))
                head.setToolTip(f'Δ по трассе «{name}»:\n'
                                '• горизонтальный маркер — ширина по его уровню вокруг максимума;\n'
                                '• вертикальный маркер — разница с опорным маркером')
        table.verticalHeader().setVisible(False)
        table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        table.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)
        table.setAlternatingRowColors(True)
        table.setWordWrap(False)
        table.setMaximumHeight(150)
        header = table.horizontalHeader()
        header.setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        header.setSectionResizeMode(0, QtWidgets.QHeaderView.ResizeToContents)
        header.setSectionResizeMode(1, QtWidgets.QHeaderView.ResizeToContents)
        header.setDefaultAlignment(QtCore.Qt.AlignCenter)
        table.setToolTip('Строка на маркер. Клик по строке подсвечивает линию на графике,\n'
                         'Del — убрать маркер')
        table.installEventFilter(self)
        table.itemSelectionChanged.connect(self._sync_active_cursor)
        table.hide()             # пока маркеров нет, место не занимаем
        self.marker_table = table
        return table

    def eventFilter(self, obj, event):
        """Del в таблице убирает выбранный маркер.

        Попасть правым кликом в линию шириной 8 px на графике трудно, а строку
        выбрать легко.
        """
        if (obj is getattr(self, 'marker_table', None)
                and event.type() == QtCore.QEvent.KeyPress
                and event.key() in (QtCore.Qt.Key_Delete, QtCore.Qt.Key_Backspace)):
            row = obj.currentRow()
            if 0 <= row < len(self._cursors):
                self._remove_cursor(self._cursors[row])
            return True
        return super().eventFilter(obj, event)

    def _cursor_title(self, cur, index):
        if cur.get('peak'):
            return 'Макс'
        kind = 'V' if cur['kind'] == 'v' else 'H'
        same = [c for c in self._cursors[:index]
                if c['kind'] == cur['kind'] and not c.get('peak')]
        return f'{kind}{len(same) + 1}'

    def _cursor_index(self, cur):
        """Номер маркера в списке (по тождеству: записи — словари)."""
        for index, other in enumerate(self._cursors):
            if other is cur:
                return index
        return -1

    def _ref_cursor(self):
        """Опорный маркер для Δ вертикальных: маркер максимума, иначе первый V."""
        if self._peak_cursor is not None and self._cursor_index(self._peak_cursor) >= 0:
            return self._peak_cursor
        for cur in self._cursors:
            if cur['kind'] == 'v':
                return cur
        return None

    def _trace_by_name(self, name):
        for trace in self._traces:
            if trace['name'] == name:
                return trace
        return None

    def _trace_value(self, trace, angle):
        """Значение трассы под углом (None — нет данных или обрезанный участок)."""
        if trace['x'].size == 0:
            return None
        y = float(np.interp(angle, trace['x'], self._disp(trace)))
        return y if np.isfinite(y) else None

    def _lobe_width(self, trace, level):
        """(ширина, все пересечения) трассы с уровнем.

        Ширина считается по паре пересечений ВОКРУГ максимума трассы. Когда
        уровень маркера ниже боковых лепестков, пересечений больше двух, и
        «крайнее минус крайнее» дало бы не ширину луча, а размах по всем
        лепесткам — то есть заведомо неверную ширину.
        """
        disp = self._disp(trace)
        crossings = self._crossings(trace['x'], disp, level)
        if len(crossings) < 2:
            return None, crossings
        finite = np.where(np.isfinite(disp), disp, -np.inf)
        x_max = float(trace['x'][int(np.argmax(finite))])
        left = [c for c in crossings if c <= x_max]
        right = [c for c in crossings if c >= x_max]
        if left and right:
            return min(right) - max(left), crossings
        return max(crossings) - min(crossings), crossings

    # Тексты-заглушки: их красим приглушённо, чтобы цифры не тонули среди прочерков.
    EMPTY = '—'
    REF_MARK = 'опора'
    NO_CROSS = 'нет пересеч.'
    _MUTED = (EMPTY, REF_MARK, NO_CROSS)

    @staticmethod
    def _fmt_value(value, unit, sign=False):
        """Значение с единицей: у градуса пробела перед знаком не бывает."""
        number = f'{value:+.2f}' if sign else f'{value:.2f}'
        return f'{number}{unit}' if unit == '°' else f'{number} {unit}'

    def _crossings_text(self, crossings):
        if not crossings:
            return self.NO_CROSS
        if len(crossings) <= 3:
            return ' / '.join(f'{x:.2f}°' for x in crossings)
        # больше трёх в ячейку не влезает — полный список уходит в подсказку
        return f'{crossings[0]:.2f}° … {crossings[-1]:.2f}° ({len(crossings)} шт.)'

    def _cursor_hint(self, cur, title, value):
        if cur.get('peak'):
            what = f'маркер максимума, {value:.3f}° (магнитится к локальным максимумам)'
        elif cur['kind'] == 'v':
            what = f'вертикальный маркер, {value:.3f}°'
        else:
            what = f'горизонтальный маркер, уровень {value:.2f} {self._left_unit}'
        return (f'{title} — {what}.\n'
                'Выделить строку — линия на графике станет сплошной и толще. Del — убрать маркер')

    def _vcursor_cells(self, cur, title, angle):
        """Ячейки вертикального маркера: значение трассы и Δ к опорному маркеру."""
        ref = self._ref_cursor()
        ref_title, ref_angle = '', None
        if ref is not None and ref is not cur:
            ref_title = self._cursor_title(ref, self._cursor_index(ref))
            ref_angle = float(ref['line'].value())

        cells = {}
        for name, _color, axis, _dashed in self._trace_defs:
            trace = self._trace_by_name(name)
            if trace is None or not trace['visible'] or trace['x'].size == 0:
                hidden = (self.EMPTY, f'«{name}»: трасса скрыта или не рассчитана')
                cells[('trace', name)] = hidden
                if axis == 'L':
                    cells[('delta', name)] = hidden
                continue

            value = self._trace_value(trace, angle)
            if value is None:
                cells[('trace', name)] = (self.EMPTY,
                                          f'«{name}»: в {angle:.3f}° значения нет '
                                          '(обрезанный нефизичный участок ДН)')
            else:
                cells[('trace', name)] = (self._fmt_value(value, trace['unit']),
                                          f'«{name}» в {angle:.3f}°')
            if axis != 'L':
                continue

            if ref_angle is None:
                cells[('delta', name)] = (
                    self.REF_MARK,
                    'Опорный маркер: от него считаются Δ остальных вертикальных.\n'
                    'Опорным берётся маркер максимума, а если его нет — первый вертикальный')
                continue
            ref_value = self._trace_value(trace, ref_angle)
            if value is None or ref_value is None:
                cells[('delta', name)] = (self.EMPTY,
                                          f'«{name}»: на одном из маркеров значения нет')
            else:
                diff = self._fmt_value(value - ref_value, trace['unit'], sign=True)
                cells[('delta', name)] = (
                    diff, f'«{name}»: {title} − {ref_title} = {diff}')
        return cells

    def _hcursor_cells(self, level):
        """Ячейки горизонтального маркера: углы пересечения и ширина по уровню."""
        cells = {}
        for name, _color, axis, _dashed in self._trace_defs:
            if axis != 'L':
                cells[('trace', name)] = (
                    self.EMPTY,
                    'Горизонтальный маркер задан по левой оси (амплитуда) —\n'
                    f'с фазовой трассой «{name}» он не пересекается')
                continue
            trace = self._trace_by_name(name)
            if trace is None or not trace['visible'] or trace['x'].size == 0:
                hidden = (self.EMPTY, f'«{name}»: трасса скрыта или не рассчитана')
                cells[('trace', name)] = hidden
                cells[('delta', name)] = hidden
                continue

            width, crossings = self._lobe_width(trace, level)
            if crossings:
                tip = (f'«{name}» пересекает уровень {level:.2f} {self._left_unit} в:\n'
                       + ', '.join(f'{x:.3f}°' for x in crossings))
            else:
                tip = f'«{name}» не доходит до уровня {level:.2f} {self._left_unit}'
            cells[('trace', name)] = (self._crossings_text(crossings), tip)

            if width is None:
                cells[('delta', name)] = (
                    self.EMPTY,
                    f'«{name}»: пересечений меньше двух — ширину считать не по чему')
            else:
                cells[('delta', name)] = (
                    f'{width:.3f}°',
                    f'«{name}»: ширина по уровню {level:.2f} {self._left_unit}\n'
                    'между пересечениями вокруг максимума трассы')
        return cells

    def _cursor_row(self, cur, index):
        """Строка таблицы: [(текст, подсказка)] в порядке self._marker_cols."""
        value = float(cur['line'].value())
        title = self._cursor_title(cur, index)
        if cur['kind'] == 'v':
            body = self._vcursor_cells(cur, title, value)
            position = (f'{value:.3f}°', f'{title}: угол маркера')
        else:
            body = self._hcursor_cells(value)
            position = (f'{value:.2f} {self._left_unit}', f'{title}: уровень маркера')

        head = (title, self._cursor_hint(cur, title, value))
        row = []
        for kind, name, _color, _dashed in self._marker_cols:
            if kind == 'title':
                row.append(head)
            elif kind == 'pos':
                row.append(position)
            else:
                row.append(body.get((kind, name), (self.EMPTY, '')))
        return row

    def _fill_marker_table(self):
        table = self.marker_table
        blocked = table.blockSignals(True)   # перезаполнение не должно сбивать выделение
        table.setRowCount(len(self._cursors))
        for row, cur in enumerate(self._cursors):
            cells = self._cursor_row(cur, row)
            for column, (kind, _name, color, _dashed) in enumerate(self._marker_cols):
                text, tip = cells[column]
                item = QtWidgets.QTableWidgetItem(text)
                item.setToolTip(tip)
                item.setTextAlignment(QtCore.Qt.AlignCenter)
                if kind == 'title':
                    # Тот же чип, что и на линии: по нему строка находится глазами
                    item.setIcon(self._swatch(cur['color'], dashed=True))
                    item.setTextAlignment(QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter)
                    item.setForeground(QtGui.QBrush(QtGui.QColor(cur['color'])))
                    font = item.font()
                    font.setBold(True)
                    item.setFont(font)
                elif text in self._MUTED:
                    item.setForeground(QtGui.QBrush(QtGui.QColor(PLOT_AXIS_MUTED)))
                elif kind in ('trace', 'delta'):
                    item.setForeground(QtGui.QBrush(QtGui.QColor(color)))
                    if kind == 'delta':
                        font = item.font()
                        font.setBold(True)     # ширина/разница — то, ради чего маркер и ставят
                        item.setFont(font)
                table.setItem(row, column, item)
        table.resizeRowsToContents()
        table.setVisible(bool(self._cursors))
        self._restore_marker_selection()
        table.blockSignals(blocked)
        self._sync_marker_labels()
        self._apply_marker_pens()

    # --------------------------------------------- связь «строка ↔ линия»
    def _restore_marker_selection(self):
        """Вернуть выделение на ту же линию после перезаполнения таблицы."""
        table = self.marker_table
        row = self._cursor_index(self._active_cursor) if self._active_cursor is not None else -1
        if row < 0:
            self._active_cursor = None
            table.clearSelection()
        elif table.currentRow() != row:
            table.selectRow(row)

    def _sync_active_cursor(self):
        row = self.marker_table.currentRow()
        self._active_cursor = self._cursors[row] if 0 <= row < len(self._cursors) else None
        self._apply_marker_pens()

    def _select_cursor_row(self, cur):
        """Клик по линии на графике — выделить её строку (обратная связь к таблице)."""
        row = self._cursor_index(cur)
        if row >= 0:
            self.marker_table.selectRow(row)

    def _apply_marker_pens(self):
        """Выделенная строка = подсвеченная линия: сплошная и толще.

        Подписи на линиях отвечают на «где какой маркер» при беглом взгляде,
        подсветка — когда маркеров полдюжины и подписи стоят рядом.
        """
        for cur in self._cursors:
            active = cur is self._active_cursor
            if cur.get('active') == active:
                continue      # таблица перезаполняется на каждый сдвиг линии
            cur['active'] = active
            base = PEAK_WIDTH if cur.get('peak') else MARKER_WIDTH
            cur['line'].setPen(pg.mkPen(
                cur['color'], width=base + (1.4 if active else 0),
                style=QtCore.Qt.SolidLine if active else QtCore.Qt.DashLine))

    def _sync_marker_labels(self):
        """Подпись на самой линии: «H1 · -3.00 дБ».

        Имена в таблице бесполезны, пока линии на графике безымянные: два
        вертикальных маркера выглядят одинаково, и какой из них V2 — непонятно.
        Имена зависят от порядка, поэтому пересобираются при каждом обновлении.
        """
        for index, cur in enumerate(self._cursors):
            label = getattr(cur['line'], 'label', None)
            if label is None:
                continue
            title = self._cursor_title(cur, index)
            unit = '°' if cur['kind'] == 'v' else f' {self._left_unit}'
            fmt = f'{title} · {{value:0.2f}}{unit}'
            if cur.get('label_fmt') == fmt:
                continue      # значение подпись пересчитывает сама, при сдвиге линии
            cur['label_fmt'] = fmt
            label.setFormat(fmt)
            if not label.isVisible():
                # setFormat перерисовывает подпись только у видимой линии,
                # а таблица заполняется и до показа окна
                label.setText(fmt.format(value=float(cur['line'].value())))

    def _build_plane_bar(self):
        """Тумблеры видимости трасс — в правом верхнем углу над графиком."""
        bar = QtWidgets.QHBoxLayout()
        bar.setContentsMargins(0, 0, 4, 0)
        bar.setSpacing(4)
        bar.addStretch(1)
        bar.addWidget(QtWidgets.QLabel('Трассы:'))
        for name, color, _axis, _dashed in self._trace_defs:
            btn = QtWidgets.QToolButton()
            btn.setText(name)
            btn.setCheckable(True)
            btn.setChecked(True)
            btn.setToolTip(f'Показать/скрыть «{name}»')
            btn.setStyleSheet(f'QToolButton:checked {{ color: {color}; font-weight: 600; }}')
            btn.clicked.connect(lambda _=False, n=name: self._toggle_trace(n))
            reserve_bold_width(btn)   # выбранный тумблер жирный — иначе текст обрежется
            bar.addWidget(btn)
            self._vis_btns[name] = btn
        return bar

    # --------------------------------------------------------------- данные
    def _disp(self, trace):
        if trace['amp'] and self._normalize:
            return trace['y_raw'] - trace['max_val']
        return trace['y_raw']

    def _visible_traces(self):
        return [t for t in self._traces if t['visible'] and t['x'].size]

    def _visible_amp_traces(self):
        return [t for t in self._traces if t['visible'] and t['amp'] and t['x'].size]

    def _toggle_trace(self, name):
        btn = self._vis_btns.get(name)
        for trace in self._traces:
            if trace['name'] == name:
                trace['visible'] = btn.isChecked() if btn else True
                if trace['visible'] and trace['x'].size:
                    trace['curve'].setData(trace['x'], self._disp(trace))
                    trace['curve'].show()
                else:
                    trace['curve'].hide()
        self._update_intersections()

    def set_data(self, traces_data):
        """traces_data: список (x, y_raw, max_val) в порядке trace_defs."""
        for trace, (x, y_raw, max_val) in zip(self._traces, traces_data):
            trace['x'] = np.asarray(x, dtype=float)
            trace['y_raw'] = np.asarray(y_raw, dtype=float)
            trace['max_val'] = float(max_val)
            if trace['x'].size and trace['visible']:
                trace['curve'].setData(trace['x'], self._disp(trace))
                trace['curve'].show()
            else:
                trace['curve'].setData([], [])
        self._refresh_mask()
        self._update_intersections()

    def clear_data(self):
        self.set_data([(np.array([]), np.array([]), 0.0) for _ in self._traces])

    def hold_current(self, label):
        self._overlay_count += 1
        color = OVERLAY_COLORS[(self._overlay_count - 1) % len(OVERLAY_COLORS)]
        for trace in self._traces:
            if trace['x'].size == 0:
                continue
            name = f'{label} · {trace["name"]}'
            curve = pg.PlotDataItem(trace['x'].copy(), self._disp(trace).copy(),
                                    pen=pg.mkPen(color, width=1, style=QtCore.Qt.DotLine), name=name,
                                    connect='finite')
            trace['vb'].addItem(curve)
            self.legend.addItem(curve, name)
            self._overlays.append([curve, trace['vb'], name])

    def clear_overlays(self):
        for curve, vb, name in self._overlays:
            vb.removeItem(curve)
            try:
                self.legend.removeItem(name)
            except Exception:
                pass
        self._overlays = []

    # ---------------------------------------------------------- управление
    def autoscale(self):
        self._pi.enableAutoRange(x=True)
        self.vbL.enableAutoRange(y=True)
        self.vbR.enableAutoRange(y=True)

    def reset_view(self):
        self._pi.enableAutoRange(x=True)
        self.vbL.setYRange(*self._left_range)
        self.vbR.setYRange(*self._right_range)

    def _toggle_norm(self):
        self._normalize = self.norm_btn.isChecked()
        for trace in self._traces:
            if trace['amp'] and trace['x'].size and trace['visible']:
                trace['curve'].setData(trace['x'], self._disp(trace))
        self._refresh_mask()
        self._update_intersections()

    # ----------------------------------------------------- маска УБЛ
    def set_sll_mask(self, db_below):
        """Включить маску (db_below — предел в дБ от максимума, <0) или None — выкл."""
        self._mask_db = db_below
        self._refresh_mask()

    def _mask_level(self):
        amps = self._visible_amp_traces()
        if not amps or self._mask_db is None:
            return None
        tops = [float(np.nanmax(d)) for d in (self._disp(t) for t in amps)
                if np.any(np.isfinite(d))]
        if not tops:
            return None
        return max(tops) + self._mask_db

    def _refresh_mask(self):
        level = self._mask_level()
        if level is None:
            if self._mask_line is not None:
                self.plot.removeItem(self._mask_line)
                self._mask_line = None
            return
        if self._mask_line is None:
            self._mask_line = pg.InfiniteLine(
                angle=0, movable=False,
                pen=pg.mkPen(MASK_COLOR, width=1.6, style=QtCore.Qt.DashDotLine),
                label='Маска УБЛ', labelOpts={'position': 0.04, 'color': MASK_COLOR})
            self._mask_line.setZValue(5)
            self.plot.addItem(self._mask_line)
        self._mask_line.setValue(level)

    # ------------------------------------------------------------ курсоры
    @staticmethod
    def _crossings(x, y, level):
        """Точки пересечения ломаной (x, y) с горизонтальным уровнем (интерп.)."""
        x = np.asarray(x, dtype=float)
        y = np.asarray(y, dtype=float)
        d = y - level
        res = []
        for i in range(d.size - 1):
            d0, d1 = d[i], d[i + 1]
            if d0 == 0.0:
                res.append(float(x[i]))
            if (d0 < 0 < d1) or (d1 < 0 < d0):
                t = d0 / (d0 - d1)
                res.append(float(x[i] + t * (x[i + 1] - x[i])))
        if d.size and d[-1] == 0.0:
            res.append(float(x[-1]))
        return res

    def _make_dots(self, vb, color):
        dots = pg.ScatterPlotItem(size=10, symbol='o',
                                  brush=pg.mkBrush(color), pen=pg.mkPen('w', width=1))
        dots.setZValue(40)
        vb.addItem(dots, ignoreBounds=True)
        return dots

    def _new_cursor_record(self, line, kind, peak=False):
        if peak:
            color = PEAK_COLOR
        else:
            color = MARKER_V_COLOR if kind == 'v' else MARKER_H_COLOR
        return {
            'line': line, 'kind': kind, 'peak': peak, 'color': color,
            'dotsL': self._make_dots(self.vbL, color),
            'dotsR': self._make_dots(self.vbR, color),
            'labelsL': [], 'labelsR': [],
        }

    @staticmethod
    def _marker_label_opts(color, position):
        """Подпись маркера — на плашке: над трассами голый текст не читался."""
        return {'position': position, 'color': color, 'border': pg.mkPen(color, width=1),
                'fill': pg.mkBrush(255, 255, 255, 235)}

    def _make_vcursor(self, x, peak=False):
        color = PEAK_COLOR if peak else MARKER_V_COLOR
        width = PEAK_WIDTH if peak else MARKER_WIDTH
        line = pg.InfiniteLine(pos=x, angle=90, movable=True,
                               pen=pg.mkPen(color, width=width, style=QtCore.Qt.DashLine),
                               label='{value:0.2f}°',
                               labelOpts=self._marker_label_opts(color, 0.97))
        line.sigPositionChanged.connect(self._update_intersections)
        if peak:
            # Маркер поиска максимума «магнитится» к локальным максимумам при сдвиге.
            line.sigDragged.connect(self._snap_peak_to_nearest)
        self.plot.addItem(line)
        cur = self._new_cursor_record(line, 'v', peak=peak)
        self._cursors.append(cur)
        # тянут линию — подсвечиваем её строку: клик по графику при перетаскивании
        # не приходит, и таблица иначе показывала бы выделенным другой маркер
        line.sigDragged.connect(lambda _line, c=cur: self._select_cursor_row(c))
        return cur

    def _add_vcursor(self):
        x0 = 0.0
        vis = self._visible_traces()
        if vis:
            x0 = float(vis[0]['x'][vis[0]['x'].size // 2])
        self._make_vcursor(x0)
        self._update_intersections()

    def _add_hcursor(self):
        y0 = -3.0
        vis = self._visible_amp_traces()
        if vis:
            d0 = self._disp(vis[0])
            if np.any(np.isfinite(d0)):
                y0 = float(np.nanmax(d0)) - 3.0
        line = pg.InfiniteLine(pos=y0, angle=0, movable=True,
                               pen=pg.mkPen(MARKER_H_COLOR, width=MARKER_WIDTH, style=QtCore.Qt.DashLine),
                               label='{value:0.2f}',
                               labelOpts=self._marker_label_opts(MARKER_H_COLOR, 0.95))
        line.sigPositionChanged.connect(self._update_intersections)
        self.plot.addItem(line)
        cur = self._new_cursor_record(line, 'h')
        self._cursors.append(cur)
        line.sigDragged.connect(lambda _line, c=cur: self._select_cursor_row(c))
        self._update_intersections()

    # ---------------------------------------------------- поиск максимумов
    # Только по амплитудным (левым) трассам. Чтобы искать в одной плоскости —
    # скройте вторую амплитудную трассу тумблером справа сверху.
    def _peak_vcursor(self, x):
        if self._peak_cursor is None or self._peak_cursor not in self._cursors:
            self._peak_cursor = self._make_vcursor(x, peak=True)
        else:
            self._peak_cursor['line'].setValue(x)
        self._update_intersections()

    def _local_max_angles(self):
        """Отсортированные углы (°) всех локальных максимумов видимых амплитуд."""
        angles = []
        for trace in self._visible_amp_traces():
            for p in find_peak_indices(self._disp(trace)):
                angles.append(float(trace['x'][p]))
        if not angles:
            return []
        return sorted(set(round(a, 6) for a in angles))

    def _snap_peak_to_nearest(self):
        """Притянуть маркер-максимум к ближайшему локальному максимуму (магнит)."""
        if self._peak_cursor is None or self._peak_cursor not in self._cursors:
            return
        angles = self._local_max_angles()
        if not angles:
            return
        x = float(self._peak_cursor['line'].value())
        nearest = min(angles, key=lambda a: abs(a - x))
        if abs(nearest - x) > 1e-9:
            self._peak_cursor['line'].setValue(nearest)

    def _peak_cursor_x(self):
        if self._peak_cursor is not None and self._peak_cursor in self._cursors:
            return float(self._peak_cursor['line'].value())
        return -np.inf

    def _marker_to_max(self):
        best_x, best_y = None, None
        for trace in self._visible_amp_traces():
            disp = self._disp(trace)
            if not np.any(np.isfinite(disp)):
                continue
            i = int(np.nanargmax(disp))
            if best_y is None or disp[i] > best_y:
                best_y, best_x = float(disp[i]), float(trace['x'][i])
        if best_x is not None:
            self._peak_vcursor(best_x)

    def _next_local_max(self):
        angles = self._local_max_angles()
        if not angles:
            return
        cur_x = self._peak_cursor_x()
        nxt = [a for a in angles if a > cur_x + 1e-6]
        self._peak_vcursor(nxt[0] if nxt else angles[0])

    # -------------------------------------- удаление/правка линий (ПКМ / 2×клик)
    def _remove_cursor(self, cur):
        self.plot.removeItem(cur['line'])
        self.vbL.removeItem(cur['dotsL'])
        self.vbR.removeItem(cur['dotsR'])
        for lab in cur['labelsL']:
            self.vbL.removeItem(lab)
        for lab in cur['labelsR']:
            self.vbR.removeItem(lab)
        row = self._cursor_index(cur)
        if row >= 0:
            self._cursors.pop(row)
        if cur is self._peak_cursor:
            self._peak_cursor = None
        if cur is self._active_cursor:
            self._active_cursor = None
        self._update_intersections()

    def _cursor_near(self, scene_pos, thresh_px=8):
        mp = self.vbL.mapSceneToView(scene_pos)
        try:
            xpix, ypix = self.vbL.viewPixelSize()
        except Exception:
            return None
        best, best_d = None, thresh_px
        for cur in self._cursors:
            v = float(cur['line'].value())
            d = abs(mp.x() - v) / xpix if cur['kind'] == 'v' else abs(mp.y() - v) / ypix
            if d < best_d:
                best_d, best = d, cur
        return best

    def _on_scene_click(self, ev):
        try:
            cur = self._cursor_near(ev.scenePos())
            if cur is not None:
                if ev.button() == QtCore.Qt.RightButton:
                    self._remove_cursor(cur)
                    ev.accept()
                elif ev.button() == QtCore.Qt.LeftButton:
                    self._select_cursor_row(cur)   # клик по линии подсвечивает её строку
                    if ev.double():
                        self._edit_cursor(cur)
                        ev.accept()
                return
            # Клик не по линии: двойной ЛКМ по полю графика — автомасштаб.
            if (ev.button() == QtCore.Qt.LeftButton and ev.double()
                    and self.plot.sceneBoundingRect().contains(ev.scenePos())):
                self.autoscale()
                ev.accept()
        except Exception:
            logger.exception('Ошибка обработки клика по графику')

    def _edit_cursor(self, cur):
        line = cur['line']
        if cur['kind'] == 'v':
            title, label = 'Вертикальный маркер', 'Угол, °:'
        else:
            title, label = 'Горизонтальный маркер', f'Уровень, {self._left_unit}:'
        val, ok = QtWidgets.QInputDialog.getDouble(self, title, label, float(line.value()),
                                                   -1.0e9, 1.0e9, 3)
        if ok:
            line.setValue(val)

    # Точки и подписи НА графике. Числа для чтения — в таблице маркеров
    # (см. _cursor_row), поэтому здесь только координаты и краткая подпись.
    def _v_points(self, x0):
        out = []
        for trace in self._visible_traces():
            y = float(np.interp(x0, trace['x'], self._disp(trace)))
            if not np.isfinite(y):     # обрезанный участок ДН — нет значения
                continue
            out.append({'x': x0, 'y': y, 'axis': trace['axis'], 'color': trace['color'],
                        'label': f'{y:.2f} {trace["unit"]}'})
        return out

    def _h_points(self, level):
        out = []
        for trace in self._visible_amp_traces():
            for xv in self._crossings(trace['x'], self._disp(trace), level):
                out.append({'x': xv, 'y': level, 'axis': 'L',
                            'color': trace['color'], 'label': f'{xv:.2f}°'})
        return out

    def _sync_labels(self, store, points, vb):
        while len(store) < len(points):
            # Крупный шрифт + компактная непрозрачная плашка, чтобы цифры
            # читались и не сливались с линией.
            t = pg.TextItem(anchor=(0.5, 1.4),
                            fill=pg.mkBrush(255, 255, 255, 230),
                            border=pg.mkPen('#cbd5e1'))
            t.setFont(self._label_font)
            t.setZValue(41)
            vb.addItem(t, ignoreBounds=True)
            store.append(t)
        for i, lab in enumerate(store):
            if i < len(points):
                p = points[i]
                lab.setText(p['label'])
                lab.setColor(p['color'])
                lab.setPos(p['x'], p['y'])
                lab.show()
            else:
                lab.hide()

    def _render_intersection(self, cur, points):
        left = [p for p in points if p['axis'] == 'L']
        right = [p for p in points if p['axis'] == 'R']
        cur['dotsL'].setData([p['x'] for p in left], [p['y'] for p in left])
        cur['dotsR'].setData([p['x'] for p in right], [p['y'] for p in right])
        self._sync_labels(cur['labelsL'], left, self.vbL)
        self._sync_labels(cur['labelsR'], right, self.vbR)

    def _update_intersections(self):
        """Точки и подписи на графике + строки в таблице маркеров."""
        for cur in self._cursors:
            val = float(cur['line'].value())
            points = self._v_points(val) if cur['kind'] == 'v' else self._h_points(val)
            self._render_intersection(cur, points)
        self._fill_marker_table()

    def clear_annotations(self):
        for cur in list(self._cursors):
            self._remove_cursor(cur)
        self._cursors = []
        self._active_cursor = None
        self._fill_marker_table()
        self._update_readout()

    # ------------------------------------------------- наведение (значение точки)
    def _nearest_point(self, scene_pos):
        """Ближайшая точка среди видимых трасс (пиксельное расстояние по своей оси)."""
        best, best_d = None, None
        for ti, trace in enumerate(self._traces):
            if not trace['visible'] or trace['x'].size == 0:
                continue
            vb = trace['vb']
            mp = vb.mapSceneToView(scene_pos)
            try:
                xpix, ypix = vb.viewPixelSize()
            except Exception:
                continue
            if xpix == 0 or ypix == 0:
                continue
            disp = self._disp(trace)
            d = ((trace['x'] - mp.x()) / xpix) ** 2 + ((disp - mp.y()) / ypix) ** 2
            i = int(np.argmin(d))
            if best_d is None or d[i] < best_d:
                best_d, best = float(d[i]), (ti, i)
        return best

    def _point_text(self, trace, xv, yv):
        return f'{trace["name"]}: {xv:.2f}°  {yv:.2f} {trace["unit"]}'

    def _hide_hover(self):
        for h in self._hover.values():
            h['dot'].setData([], [])
            h['lab'].hide()
        self.vline.hide()
        self._hover_text = ''

    def _on_mouse_move(self, evt):
        try:
            pos = evt[0]
            if not self.plot.sceneBoundingRect().contains(pos):
                self._hide_hover()
                self._update_readout()
                return
            hit = self._nearest_point(pos)
            if hit is None:
                return
            ti, idx = hit
            trace = self._traces[ti]
            xv = float(trace['x'][idx])
            yv = float(self._disp(trace)[idx])
            for key, h in self._hover.items():
                if key == trace['axis']:
                    h['dot'].setData([xv], [yv])
                    h['dot'].setBrush(pg.mkBrush(trace['color']))
                    h['lab'].setText(self._point_text(trace, xv, yv))
                    h['lab'].setColor(trace['color'])
                    h['lab'].setPos(xv, yv)
                    h['lab'].show()
                else:
                    h['dot'].setData([], [])
                    h['lab'].hide()
            self.vline.setPos(xv)
            self.vline.show()
            self._hover_text = self._point_text(trace, xv, yv)
            self._update_readout()
        except Exception:
            pass

    def _update_readout(self):
        """Строка под графиком — только значение под курсором.

        Значения маркеров ушли в таблицу: в одну строку они не помещались.
        """
        self.readout.setText('▶ ' + self._hover_text if self._hover_text else ' ')

    # ------------------------------------------------------------ экспорт
    def _export_png(self):
        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, 'Сохранить график', 'far_field.png', 'PNG (*.png)')
        if not path:
            return
        try:
            pg.exporters.ImageExporter(self._pi).export(path)
            logger.info(f'График сохранён: {path}')
        except Exception as exc:
            logger.error(f'Ошибка экспорта PNG: {exc}')
            QtWidgets.QMessageBox.critical(self, 'Ошибка экспорта', str(exc))

    def _csv_columns(self):
        """Колонки для CSV: на каждую сетку углов свой столбец угла и её трассы.

        Az и El считаются по разным сеткам, поэтому столбцов угла может быть два.
        """
        groups = []                      # [(сетка_углов, [трассы])]
        for trace in self._traces:
            if trace['x'].size == 0:
                continue
            for grid, members in groups:
                if grid.size == trace['x'].size and np.allclose(grid, trace['x']):
                    members.append(trace)
                    break
            else:
                groups.append((trace['x'], [trace]))

        columns = []
        for grid, members in groups:
            plane = members[0]['name'].split()[0]    # «Az ампл.» -> «Az»
            columns.append((f'Угол {plane}, °', grid))
            for trace in members:
                columns.append((f'{trace["name"]}, {trace["unit"]}', self._disp(trace)))
        return columns

    def _export_csv(self):
        if not any(t['x'].size for t in self._traces):
            return
        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, 'Сохранить данные', 'far_field.csv', 'CSV (*.csv)')
        if not path:
            return
        columns = self._csv_columns()
        try:
            # Колонками, а не строками: у Excel предел 16384 столбца, и при мелком
            # шаге по углу (±90° с шагом 0.005 — 36001 точка) хвост трассы молча
            # пропадал. Плюс по колонкам строится график и читается pandas.
            #
            # utf-8-sig, а не utf-8: без BOM Excel на русской Windows читает файл
            # как CP1251 и подписи столбцов превращаются в «CfPiPsP»».
            with open(path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f, delimiter=';')
                writer.writerow([title for title, _ in columns])
                depth = max(len(values) for _, values in columns)
                for i in range(depth):
                    writer.writerow([_csv_num(values[i]) if i < len(values) else ''
                                     for _, values in columns])
            logger.info(f'Данные сохранены: {path} ({len(columns)} столбцов)')
        except Exception as exc:
            logger.error(f'Ошибка экспорта CSV: {exc}')
            QtWidgets.QMessageBox.critical(self, 'Ошибка экспорта', str(exc))


# ==================================================================== Окно
class FarFieldWindow(QtWidgets.QMainWindow):
    """Главное окно утилиты.

    QMainWindow, а не QDialog: шапка из полутора десятков виджетов не влезала
    в узкое окно и молча обрезалась, а полоса прогресса, появляясь и исчезая,
    дёргала всю компоновку. Тулбар сам прячет лишнее в меню переполнения,
    статусная строка держит место под прогресс постоянно.
    """

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle('Дальняя зона (расчёт ДН)')
        self.setMinimumSize(960, 600)
        self._resize_to_screen(1300, 820)
        self.setAcceptDrops(True)

        self._result = None
        self._folder = None
        self._single_file_mode = False
        self._beams = []
        self._freqs = []
        self._computed_beams = []
        self._computed_freqs = []
        self._view = 'near'         # 'near' — ближнее поле, 'far' — сечения ДН
        self._view_beams = []       # что сейчас в комбобоксах (зависит от вида)
        self._view_freqs = []
        self._beam_order = self._load_beam_order()
        self._cache = {}
        self._params = self._load_saved_params()
        self._default_step = (1.0, 1.0)
        self._az_deg = None
        self._el_deg = None
        self._updating = False
        self._thread = None
        self._worker = None
        self._load_thread = None
        self._load_worker = None
        self._loading = None        # (kind, path) текущей загрузки
        self._pending_load = None   # итог загрузки, применяется после остановки потока
        self._closing = False       # окно ждёт остановки потоков, чтобы закрыться

        self._build_ui()
        self._restore_window_state()

    # ------------------------------------------------ запоминание параметров
    @staticmethod
    def _settings():
        return QtCore.QSettings('FarZone', 'FarField')

    _PARAM_KEYS = ('az_from', 'az_to', 'az_step', 'el_from', 'el_to', 'el_step', 'dx', 'dy')

    def _load_saved_params(self):
        """Прочитать последние введённые параметры пересчёта (между сессиями)."""
        s = self._settings()
        s.beginGroup('far_field/params')
        params = {}
        try:
            for k in self._PARAM_KEYS:
                v = s.value(k)
                if v is not None:
                    params[k] = float(v)
            pad = s.value('pad')
            if pad is not None:
                params['pad'] = int(pad)
            for k in ('interp', 'abs_mode'):
                v = s.value(k)
                if v is not None:
                    params[k] = str(v)
        finally:
            s.endGroup()
        return params or None

    def _save_params(self, params):
        if not params:
            return
        s = self._settings()
        s.beginGroup('far_field/params')
        try:
            for k in self._PARAM_KEYS:
                if k in params:
                    s.setValue(k, float(params[k]))
            if 'pad' in params:
                s.setValue('pad', int(params['pad']))
            for k in ('interp', 'abs_mode'):
                if k in params:
                    s.setValue(k, str(params[k]))
        finally:
            s.endGroup()
        s.sync()

    # ----------------------------------- запоминание окна и последней папки
    @staticmethod
    def _available_area():
        """Рабочая область экрана (без панели задач) или None."""
        screen = QtGui.QGuiApplication.primaryScreen()
        return screen.availableGeometry() if screen is not None else None

    def _resize_to_screen(self, width, height):
        """Стартовый размер, который заведомо влезает в экран.

        Прежние 1300×820 не помещались на обычном ноутбуке 1920×1080 при
        масштабе 150 %: логически это 1280×720, и окно открывалось больше
        рабочей области (а прежний минимум 1080×700 не давал его ужать).
        """
        area = self._available_area()
        if area is not None:
            width = min(width, area.width() - 40)
            height = min(height, area.height() - 60)
        self.resize(max(width, 640), max(height, 480))

    def _fit_into_screen(self):
        """Ужать окно до экрана — геометрия могла сохраниться с монитора крупнее."""
        area = self._available_area()
        if area is None:
            return
        size = self.size()
        width = min(size.width(), area.width())
        height = min(size.height(), area.height())
        if (width, height) != (size.width(), size.height()):
            self.resize(width, height)
        if not area.intersects(self.frameGeometry()):
            self.move(area.topLeft())

    def _restore_window_state(self):
        settings = self._settings()
        geo = settings.value('far_field/geometry')
        if geo is not None:
            try:
                self.restoreGeometry(geo)
            except Exception:
                pass
        # Ширина панели метрик — вещь личная: кому-то нужны полные подписи,
        # кому-то максимум графика. Запоминаем вместе с геометрией окна.
        sizes = settings.value('far_field/splitter')
        if sizes is not None:
            try:
                self._splitter.restoreState(sizes)
            except Exception:
                pass
        self._fit_into_screen()

    def _save_window_state(self):
        settings = self._settings()
        settings.setValue('far_field/geometry', self.saveGeometry())
        settings.setValue('far_field/splitter', self._splitter.saveState())

    def _last_folder(self):
        return self._settings().value('far_field/last_folder', '') or ''

    def _set_last_folder(self, folder):
        self._settings().setValue('far_field/last_folder', folder)

    _BEAM_ORDERS = (
        ('по номеру', BEAM_ORDER_NUMBER),
        ('по азимуту α', BEAM_ORDER_AZIMUTH),
        ('по углу места β', BEAM_ORDER_ELEVATION),
    )

    def _load_beam_order(self):
        """Запомненный порядок перебора лучей (между сессиями)."""
        saved = str(self._settings().value('far_field/beam_order', '') or '')
        known = {mode for _, mode in self._BEAM_ORDERS}
        return saved if saved in known else BEAM_ORDER_NUMBER

    def _build_ui(self):
        toolbar = self._build_toolbar()
        toolbar.setObjectName('mainToolBar')   # нужен для saveState()
        self.addToolBar(toolbar)

        central = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout(central)
        layout.setContentsMargins(10, 8, 10, 4)

        # Сплиттер вместо жёстких 300 px: из-за фиксированной ширины подписи
        # метрик приходилось сокращать («Задан. α/β») и переносить по словам.
        self._splitter = QtWidgets.QSplitter(QtCore.Qt.Horizontal)
        self._splitter.setChildrenCollapsible(False)
        self._splitter.setHandleWidth(8)
        self._splitter.addWidget(self._build_left_panel())
        self._splitter.addWidget(self._build_plots())
        self._splitter.setStretchFactor(0, 0)
        self._splitter.setStretchFactor(1, 1)
        self._splitter.setSizes([320, 900])
        layout.addWidget(self._splitter)

        self.setCentralWidget(central)
        self._build_status_bar()
        self._sync_view_widgets()   # стартуем на ближнем поле: метрики ДН скрыты

    def _build_toolbar(self):
        bar = QtWidgets.QToolBar('Основная панель')
        bar.setMovable(False)
        bar.setFloatable(False)
        bar.setIconSize(QtCore.QSize(ICON_SM, ICON_SM))

        self.open_btn = QtWidgets.QPushButton('Открыть папку')
        # Пока данных нет, главное действие экрана — открыть папку; после
        # загрузки акцент переезжает на «Пересчитать…» (см. _apply_result).
        self._set_primary(self.open_btn, 'folder-open', True)
        self.open_btn.setToolTip('Выбрать папку с результатами сканирования лучей (Beam№*.xlsx)')
        self.open_btn.clicked.connect(self.open_folder)
        bar.addWidget(self.open_btn)

        self.open_file_btn = QtWidgets.QPushButton('Открыть файл')
        set_button_icon(self.open_file_btn, 'file-open')
        self.open_file_btn.setToolTip('Открыть отдельный файл измерения для пересчёта '
                                      '(имя любое — важен формат внутри); '
                                      'шаг dx/dy и др. задаются вручную')
        self.open_file_btn.clicked.connect(self.open_file)
        bar.addWidget(self.open_file_btn)

        self.recalc_btn = QtWidgets.QPushButton('Пересчитать…')
        set_button_icon(self.recalc_btn, 'recalc')
        self.recalc_btn.setToolTip('Задать параметры пересчёта (диапазоны, dx/dy, лучи, частоты)\n'
                                   'и посчитать дальнюю зону')
        self.recalc_btn.clicked.connect(self.open_params)
        self.recalc_btn.setEnabled(False)
        bar.addWidget(self.recalc_btn)

        bar.addSeparator()
        self.near_btn = self._view_button('near-field', 'Ближнее поле',
                                          'Измеренное поле по апертуре (2D)', 'near')
        self.far_btn = self._view_button('far-field', 'Главные сечения ДН',
                                         'Диаграмма направленности в дальней зоне.\n'
                                         'Доступно после пересчёта', 'far')
        self.near_btn.setChecked(True)
        self.far_btn.setEnabled(False)
        bar.addWidget(self.near_btn)
        bar.addWidget(self.far_btn)

        bar.addSeparator()
        self.hold_btn = QtWidgets.QPushButton('Закрепить')
        set_button_icon(self.hold_btn, 'pin')
        self.hold_btn.setToolTip('Закрепить текущие трассы для сравнения')
        self.hold_btn.clicked.connect(self._hold_traces)
        self.hold_btn.setEnabled(False)
        bar.addWidget(self.hold_btn)

        self.clear_overlays_btn = QtWidgets.QPushButton('Очистить наложения')
        set_button_icon(self.clear_overlays_btn, 'eraser')
        self.clear_overlays_btn.setToolTip('Убрать закреплённые трассы')
        self.clear_overlays_btn.clicked.connect(self._clear_overlays)
        self.clear_overlays_btn.setEnabled(False)
        bar.addWidget(self.clear_overlays_btn)

        bar.addSeparator()
        self.beam_kind_label = QtWidgets.QLabel('Луч:')
        bar.addWidget(self.beam_kind_label)
        self.beam_prev_btn = self._nav_button('back', 'Предыдущий луч  [←]', self._beam_prev,
                                              QtGui.QKeySequence(QtCore.Qt.Key_Left))
        bar.addWidget(self.beam_prev_btn)
        self.beam_combo = QtWidgets.QComboBox()
        self.beam_combo.setMinimumWidth(90)
        # Рядом с номером луча показываются его углы — ширины под номер мало.
        self.beam_combo.setSizeAdjustPolicy(QtWidgets.QComboBox.AdjustToContents)
        self.beam_combo.currentIndexChanged.connect(self._on_selection_changed)
        bar.addWidget(self.beam_combo)
        self.beam_next_btn = self._nav_button('forward', 'Следующий луч  [→]', self._beam_next,
                                              QtGui.QKeySequence(QtCore.Qt.Key_Right))
        bar.addWidget(self.beam_next_btn)

        self.order_combo = QtWidgets.QComboBox()
        self.order_combo.setToolTip(
            'Порядок перебора лучей стрелками ←/→:\n'
            '• по номеру — как пронумерованы;\n'
            '• по азимуту α — все углы места при одном азимуте, затем следующий азимут;\n'
            '• по углу места β — все азимуты при одном угле места, затем следующий.')
        for text, mode in self._BEAM_ORDERS:
            self.order_combo.addItem(text, mode)
        self.order_combo.setCurrentIndex(max(0, self.order_combo.findData(self._beam_order)))
        self.order_combo.currentIndexChanged.connect(self._on_order_changed)
        bar.addWidget(self.order_combo)

        bar.addSeparator()
        bar.addWidget(QtWidgets.QLabel('Частота:'))
        self.freq_prev_btn = self._nav_button('back', 'Предыдущая частота  [PgUp]', self._freq_prev,
                                              QtGui.QKeySequence(QtCore.Qt.Key_PageUp))
        bar.addWidget(self.freq_prev_btn)
        self.freq_combo = QtWidgets.QComboBox()
        self.freq_combo.setMinimumWidth(110)
        self.freq_combo.currentIndexChanged.connect(self._on_selection_changed)
        bar.addWidget(self.freq_combo)
        self.freq_next_btn = self._nav_button('forward', 'Следующая частота  [PgDn]', self._freq_next,
                                              QtGui.QKeySequence(QtCore.Qt.Key_PageDown))
        bar.addWidget(self.freq_next_btn)
        return bar

    @staticmethod
    def _set_primary(button, icon_name, primary):
        """Пометить кнопку главным действием экрана (акцентная заливка).

        Заодно перекрашивает иконку: тёмно-серый глиф на индиго-заливке
        сливался бы с фоном.
        """
        button.setProperty('primary', primary)
        set_button_icon(button, icon_name,
                        color=ICON_ON_ACCENT if primary else ICON_DEFAULT)
        button.style().unpolish(button)
        button.style().polish(button)

    def _view_button(self, icon, text, tooltip, view):
        """Кнопка выбора вида (ближнее поле / дальняя зона) — как радиокнопка."""
        btn = QtWidgets.QToolButton()
        btn.setText(text)
        btn.setToolButtonStyle(QtCore.Qt.ToolButtonTextBesideIcon)
        set_button_icon(btn, icon, size=ICON_SM)
        btn.setCheckable(True)
        btn.setAutoExclusive(False)   # переключаем вручную: вид может не смениться
        btn.setToolTip(tooltip)
        btn.clicked.connect(lambda _=False, v=view: self._set_view(v))
        reserve_bold_width(btn)       # выбранная кнопка жирная — иначе текст обрежется
        return btn

    def _nav_button(self, icon, tooltip, slot, shortcut=None):
        btn = QtWidgets.QToolButton()
        btn.setProperty('navButton', True)
        set_button_icon(btn, icon, size=ICON_SM)
        if shortcut is not None:
            btn.setShortcut(shortcut)
        btn.setToolTip(tooltip)
        btn.clicked.connect(slot)
        return btn

    def _build_status_bar(self):
        """Источник данных слева, прогресс и отмена справа.

        Полоса прогресса раньше жила в компоновке окна и, появляясь, сдвигала
        график; в статусной строке место под неё есть всегда.
        """
        status = self.statusBar()
        status.setSizeGripEnabled(True)

        self.folder_label = QtWidgets.QLabel('Папка не выбрана')
        self.folder_label.setObjectName('ffFolderLabel')
        self.folder_label.setMinimumWidth(80)
        status.addWidget(self.folder_label, 1)

        self.progress = QtWidgets.QProgressBar()
        self.progress.setRange(0, 100)
        self.progress.setTextVisible(True)
        self.progress.setVisible(False)
        self.progress.setFixedWidth(340)
        status.addPermanentWidget(self.progress)

        self.cancel_btn = QtWidgets.QPushButton('Отмена')
        set_button_icon(self.cancel_btn, 'stop')
        self.cancel_btn.clicked.connect(self._cancel_current)
        self.cancel_btn.setVisible(False)
        status.addPermanentWidget(self.cancel_btn)

    def _build_left_panel(self):
        panel = QtWidgets.QWidget()
        panel.setMinimumWidth(260)
        panel.setMaximumWidth(520)
        layout = QtWidgets.QVBoxLayout(panel)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(10)

        self.near_group = QtWidgets.QGroupBox('Ближнее поле')
        n = QtWidgets.QFormLayout(self.near_group)
        n.setContentsMargins(12, 12, 12, 12)
        self.lbl_nf_angles = self._metric()
        self.lbl_nf_max = self._metric()
        self.lbl_nf_max_pos = self._metric()
        self.lbl_nf_dynamic = self._metric()
        self.lbl_nf_phase_span = self._metric()
        self.lbl_nf_points = self._metric()
        self.lbl_nf_size = self._metric()
        self.lbl_nf_grid = self._metric()
        n.addRow('Заданный α / β, °:', self.lbl_nf_angles)
        n.addRow('Максимум, дБ:', self.lbl_nf_max)
        # Где именно максимум, размах амплитуды и фазы — это уже считалось в
        # field_stats, но никуда не выводилось.
        n.addRow('Максимум в X/Y, см:', self.lbl_nf_max_pos)
        n.addRow('Размах амплитуды, дБ:', self.lbl_nf_dynamic)
        n.addRow('Размах фазы, °:', self.lbl_nf_phase_span)
        n.addRow('Измерено точек:', self.lbl_nf_points)
        n.addRow('Апертура X×Y, см:', self.lbl_nf_size)
        n.addRow('Сетка X×Y, точек:', self.lbl_nf_grid)
        layout.addWidget(self.near_group)

        # Сведения о самом скане: строки собираются по факту наличия данных,
        # поэтому у одиночного файла (без scan_params.json) группа почти пуста.
        self.scan_group = QtWidgets.QGroupBox('Параметры скана')
        self.scan_form = QtWidgets.QFormLayout(self.scan_group)
        self.scan_form.setContentsMargins(12, 12, 12, 12)
        self.scan_form.setFieldGrowthPolicy(QtWidgets.QFormLayout.AllNonFixedFieldsGrow)
        layout.addWidget(self.scan_group)

        metrics_group = QtWidgets.QGroupBox('Результаты')
        self.metrics_group = metrics_group
        m = QtWidgets.QFormLayout(metrics_group)
        m.setContentsMargins(12, 12, 12, 12)
        self.lbl_max = self._metric()
        self.lbl_pos_az = self._metric()
        self.lbl_pos_el = self._metric()
        self.lbl_sll_az = self._metric()
        self.lbl_sll_el = self._metric()
        self.lbl_bw_az = self._metric()
        self.lbl_bw_el = self._metric()
        self.lbl_phase_max = self._metric()
        self.lbl_set_angle = self._metric()
        m.addRow('Максимум, дБ:', self.lbl_max)
        m.addRow('Положение Az, °:', self.lbl_pos_az)
        m.addRow('Положение El, °:', self.lbl_pos_el)
        m.addRow('Ширина ДН Az (−3дБ), °:', self.lbl_bw_az)
        m.addRow('Ширина ДН El (−3дБ), °:', self.lbl_bw_el)
        m.addRow('УБЛ Az, дБ:', self.lbl_sll_az)
        m.addRow('УБЛ El, дБ:', self.lbl_sll_el)
        m.addRow('Фаза в макс., °:', self.lbl_phase_max)
        m.addRow('Заданный α / β, °:', self.lbl_set_angle)
        layout.addWidget(metrics_group)

        mask_group = QtWidgets.QGroupBox('Маска УБЛ')
        self.mask_group = mask_group
        mg = QtWidgets.QGridLayout(mask_group)
        mg.setContentsMargins(12, 10, 12, 10)
        self.mask_check = QtWidgets.QCheckBox('Показать маску')
        self.mask_check.setToolTip('Предел уровня боковых лепестков относительно главного максимума')
        self.mask_check.toggled.connect(self._update_mask)
        self.mask_spin = QtWidgets.QDoubleSpinBox()
        self.mask_spin.setRange(-60, -1)
        self.mask_spin.setDecimals(1)
        self.mask_spin.setSingleStep(1.0)
        self.mask_spin.setValue(-20.0)
        self.mask_spin.setSuffix(' дБ')
        self.mask_spin.valueChanged.connect(self._update_mask)
        self.mask_result = QtWidgets.QLabel('—')
        self.mask_result.setWordWrap(True)
        mg.addWidget(self.mask_check, 0, 0, 1, 2)
        mg.addWidget(QtWidgets.QLabel('Предел:'), 1, 0)
        mg.addWidget(self.mask_spin, 1, 1)
        mg.addWidget(self.mask_result, 2, 0, 1, 2)
        layout.addWidget(mask_group)

        layout.addStretch(1)

        self.export_table_btn = QtWidgets.QPushButton('Экспорт таблицы метрик…')
        set_button_icon(self.export_table_btn, 'export')
        self.export_table_btn.setToolTip('Сохранить метрики всех рассчитанных лучей и частот в Excel или CSV')
        self.export_table_btn.clicked.connect(self._export_metrics_table)
        self.export_table_btn.setEnabled(False)
        layout.addWidget(self.export_table_btn)
        return panel

    def _build_plots(self):
        self.near_panel = NearFieldPanel()
        self.plot_panel = FarFieldPlotPanel([
            ('Az ампл.', AZ_COLOR, 'L', False),
            ('El ампл.', EL_COLOR, 'L', False),
            ('Az фаза', PHASE_AZ_COLOR, 'R', True),
            ('El фаза', PHASE_EL_COLOR, 'R', True),
        ])
        self.panels = [self.plot_panel]   # закрепление трасс — только у дальней зоны

        self.stack = QtWidgets.QStackedWidget()
        self.stack.addWidget(self.near_panel)    # 0 — ближнее поле
        self.stack.addWidget(self.plot_panel)    # 1 — главные сечения ДН
        return self.stack

    @staticmethod
    def _metric():
        lbl = QtWidgets.QLabel('—')
        lbl.setProperty('metricValue', True)
        lbl.setTextInteractionFlags(QtCore.Qt.TextSelectableByMouse)
        lbl.setWordWrap(True)   # панель узкая (300 px) — длинное значение переносим, а не режем
        return lbl

    # --------------------------------------------------------------- Поток
    def _busy_warn(self):
        if self._load_worker is not None:
            QtWidgets.QMessageBox.information(self, 'Идёт загрузка',
                                              'Дождитесь окончания загрузки файлов или отмените её.')
            return True
        if self._worker is not None:
            QtWidgets.QMessageBox.information(self, 'Идёт расчёт',
                                             'Дождитесь окончания текущего расчёта или отмените его.')
            return True
        return False

    def open_folder(self):
        folder = QtWidgets.QFileDialog.getExistingDirectory(
            self, 'Выберите папку с результатами сканирования лучей', self._last_folder())
        if not folder:
            return
        self._load_folder(folder)

    def open_file(self):
        path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self, 'Выберите файл для пересчёта',
            self._last_folder(), 'Все файлы (*);;Файлы измерений (*.xlsx)')
        if not path:
            return
        self._load_file(path)

    # ------------------------------------------------------- Drag-and-drop
    @staticmethod
    def _dropped_target(mime):
        """Вернуть ('dir', path) для папки или ('file', path) для любого файла.

        Любой файл принимаем — корректность формата проверит загрузчик.
        """
        for url in mime.urls():
            if url.isLocalFile():
                path = url.toLocalFile()
                if os.path.isdir(path):
                    return 'dir', path
                if os.path.isfile(path):
                    return 'file', path
        return None

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls() and self._dropped_target(event.mimeData()):
            event.acceptProposedAction()
        else:
            event.ignore()

    def dropEvent(self, event):
        target = self._dropped_target(event.mimeData()) if event.mimeData().hasUrls() else None
        if not target:
            event.ignore()
            return
        event.acceptProposedAction()
        kind, path = target
        if kind == 'dir':
            self._load_folder(path)
        else:
            self._load_file(path)

    def _load_folder(self, folder):
        self._start_load('dir', folder)

    def _load_file(self, path):
        self._start_load('file', path)

    # ------------------------------------------------------- Загрузка в фоне
    def _start_load(self, kind, path):
        """Читать Excel в фоне: окно живое, видно прогресс и кнопку «Отмена»."""
        if self._busy_warn():
            return
        self._loading = (kind, path)
        self._pending_load = None
        self._set_busy(True)
        self.progress.setRange(0, 1000)
        self.progress.setValue(0)
        self.progress.setFormat('Подготовка…')
        name = os.path.basename(path.rstrip('/\\')) or path
        self._set_source_label(f'Загрузка: {name}', path)

        self._load_thread = QtCore.QThread()
        self._load_worker = _LoadWorker(kind, path)
        self._load_worker.moveToThread(self._load_thread)
        self._load_thread.started.connect(self._load_worker.run)
        self._load_worker.progress.connect(self._on_load_progress)
        self._load_worker.finished.connect(self._on_load_finished)
        self._load_worker.failed.connect(self._on_load_failed)
        self._load_worker.cancelled.connect(self._on_load_cancelled)
        # Итог показываем только после полной остановки потока (см. _finish_load).
        self._load_worker.done.connect(self._load_thread.quit)
        self._load_thread.finished.connect(self._on_load_thread_finished)
        self._load_thread.start()

    @QtCore.pyqtSlot(int, str)
    def _on_load_progress(self, permille, text):
        if permille < 0:
            if self.progress.maximum() != 0:
                self.progress.setRange(0, 0)   # «бегущая» полоса: доля неизвестна
            self.progress.setFormat(text)
            return
        if self.progress.maximum() != 1000:
            self.progress.setRange(0, 1000)
        self.progress.setValue(permille)
        self.progress.setFormat(f'{text}  —  %p%')

    @QtCore.pyqtSlot(object)
    def _on_load_finished(self, result):
        self._pending_load = ('ok', result)

    @QtCore.pyqtSlot(str, str)
    def _on_load_failed(self, title, message):
        self._pending_load = ('fail', (title, message))

    @QtCore.pyqtSlot()
    def _on_load_cancelled(self):
        self._pending_load = ('cancel', None)

    @QtCore.pyqtSlot()
    def _on_load_thread_finished(self):
        if self._load_worker is not None:
            self._load_worker.deleteLater()
        if self._load_thread is not None:
            self._load_thread.deleteLater()
        self._load_worker = None
        self._load_thread = None
        if self._closing:
            self._pending_load = None
            self._loading = None
            self._close_if_pending()
            return
        self._finish_load()

    def _finish_load(self):
        """Показать итог загрузки — уже после остановки потока.

        Окно параметров модальное (вложенный цикл событий), поэтому открывать
        его во время уборки потока нельзя.
        """
        pending, self._pending_load = self._pending_load, None
        loading, self._loading = self._loading, None
        self._set_busy(False)
        self.progress.setRange(0, 100)

        if not pending or not loading:
            self._restore_source_label()
            return
        outcome, payload = pending
        if outcome != 'ok':
            self._restore_source_label()
            if outcome == 'fail':
                title, message = payload
                QtWidgets.QMessageBox.warning(self, title, message)
            return

        kind, path = loading
        result = payload
        if kind == 'dir':
            self._set_last_folder(path)
            try:
                step = (float(result.get('step_x', 1.0)), float(result.get('step_y', 1.0)))
            except (TypeError, ValueError):
                step = (1.0, 1.0)
            self._apply_result(result, path, single_file=False, scan_step=step)
        else:
            self._set_last_folder(os.path.dirname(path))
            self._apply_result(result, path, single_file=True, scan_step=None)

    def _set_source_label(self, text, tooltip):
        """Подпись источника в статусной строке, с многоточием по ширине.

        QLabel режет длинное имя папки без многоточия — по обрезку не понять,
        что текст неполный. Полный путь остаётся в подсказке.
        """
        self._source_text = text
        self.folder_label.setToolTip(tooltip)
        self._elide_source_label()

    def _elide_source_label(self):
        text = getattr(self, '_source_text', '')
        metrics = QtGui.QFontMetrics(self.folder_label.font())
        width = max(self.folder_label.width(), 80)
        self.folder_label.setText(
            metrics.elidedText(text, QtCore.Qt.ElideMiddle, width))

    def resizeEvent(self, event):
        super().resizeEvent(event)
        if hasattr(self, 'folder_label'):
            self._elide_source_label()

    def _restore_source_label(self):
        """Вернуть подпись источника после неудачной или отменённой загрузки."""
        src = self._folder
        if src:
            self._set_source_label(
                os.path.basename(str(src).rstrip('/\\')) or str(src), str(src))
        else:
            self._set_source_label('Папка не выбрана', '')

    def _apply_result(self, result, source, single_file, scan_step):
        """Применить загруженный набор данных (папка или одиночный файл)."""
        self._result = result
        self._folder = source
        self._single_file_mode = single_file
        self._beams = sorted(result['data'].keys())
        self._freqs = list(result.get('freq_list') or [])
        if not self._freqs and self._beams:
            self._freqs = sorted(result['data'][self._beams[0]].keys())

        self._set_source_label(
            os.path.basename(source.rstrip('/\\')) or source, source)
        self.beam_kind_label.setText('Файл:' if single_file else 'Луч:')
        # У одиночного файла луча (а значит и углов) нет — группировать нечего.
        self.order_combo.setVisible(not single_file)
        self.recalc_btn.setEnabled(True)
        # Данные есть — теперь главное действие экрана это пересчёт.
        self._set_primary(self.open_btn, 'folder-open', False)
        self._set_primary(self.recalc_btn, 'recalc', True)

        # Новые данные — прежний расчёт к ним не относится: дальняя зона гаснет.
        self._cache = {}
        self._computed_beams = []
        self._computed_freqs = []
        self.far_btn.setEnabled(False)
        self.plot_panel.clear_data()
        self.plot_panel.clear_overlays()

        if self._params is None:
            self._params = {}
        if single_file:
            # Шаг сканера неизвестен — dx/dy вводятся вручную (берём прошлые или 1.0).
            self._default_step = (float(self._params.get('dx', 1.0)),
                                  float(self._params.get('dy', 1.0)))
        else:
            # Запомненные диапазоны/шаги/БПФ сохраняем, но dx/dy берём из шага скана.
            self._default_step = scan_step
            self._params['dx'], self._params['dy'] = self._default_step

        kind = 'файл' if single_file else 'лучей'
        logger.info(f'Загружено {kind}: {len(self._beams)}, частот: {len(self._freqs)} из {source}')
        # Сразу показываем ближнее поле; пересчёт запускает пользователь кнопкой.
        self._set_view('near', force=True)

    def open_params(self):
        if self._result is None:
            return
        defaults = dict(self._params) if self._params else {}
        defaults.setdefault('dx', self._default_step[0])
        defaults.setdefault('dy', self._default_step[1])
        dlg = FarFieldParamsDialog(self._beams, self._freqs, defaults=defaults,
                                   single_file=self._single_file_mode, parent=self)
        if dlg.exec_() != QtWidgets.QDialog.Accepted:
            return
        self._params = dlg.params()
        self._save_params(self._params)
        self._az_deg = angle_axis(self._params['az_from'], self._params['az_to'], self._params['az_step'])
        self._el_deg = angle_axis(self._params['el_from'], self._params['el_to'], self._params['el_step'])
        tasks = [(b, f) for b in dlg.selected_beams() for f in dlg.selected_freqs()]
        self._start_compute(tasks)

    def _start_compute(self, tasks):
        if not tasks:
            return
        self._set_busy(True)
        self.progress.setRange(0, len(tasks))
        self.progress.setValue(0)

        self._thread = QtCore.QThread()
        self._worker = _FarFieldWorker(self._result['data'], tasks, self._params)
        self._worker.moveToThread(self._thread)
        self._thread.started.connect(self._worker.run)
        self._worker.progress.connect(self._on_progress)
        self._worker.finished.connect(self._on_compute_finished)
        self._worker.failed.connect(self._on_compute_failed)
        # Удаляем объекты ТОЛЬКО после полной остановки потока (thread.finished),
        # иначе "QThread: Destroyed while thread is still running".
        self._worker.done.connect(self._thread.quit)
        self._thread.finished.connect(self._on_thread_finished)
        self._thread.start()

    @QtCore.pyqtSlot()
    def _on_thread_finished(self):
        if self._worker is not None:
            self._worker.deleteLater()
        if self._thread is not None:
            self._thread.deleteLater()
        self._worker = None
        self._thread = None
        self._close_if_pending()

    @QtCore.pyqtSlot(int, int, str)
    def _on_progress(self, done, total, msg):
        self.progress.setValue(done)
        self.progress.setFormat(f'{msg}  ({done}/{total})')

    @QtCore.pyqtSlot(dict)
    def _on_compute_finished(self, cache):
        try:
            self._cache = cache
            self._set_busy(False)
            if not cache:
                # Считать нечего — дальняя зона снова гаснет, остаёмся на ближнем поле.
                self._computed_beams = []
                self._computed_freqs = []
                self.far_btn.setEnabled(False)
                self._set_view('near', force=True)
                QtWidgets.QMessageBox.information(self, 'Нет результатов',
                                                  'Ни один выбранный луч/частота не содержит данных.')
                return
            self._computed_beams = sorted({b for b, _ in cache.keys()})
            self._computed_freqs = sorted({f for _, f in cache.keys()})
            # Расчёт есть — дальняя зона доступна, и сразу показываем её.
            self.far_btn.setEnabled(True)
            self._set_view('far', force=True)
        except Exception:
            logger.exception('Ошибка при отображении результатов расчёта дальней зоны')

    @QtCore.pyqtSlot(str)
    def _on_compute_failed(self, message):
        self._set_busy(False)
        QtWidgets.QMessageBox.critical(self, 'Ошибка расчёта', message)

    def _cancel_current(self):
        """Прервать то, что идёт сейчас: загрузку файлов или расчёт."""
        stopped = False
        if self._load_worker is not None:
            self._load_worker.stop()
            self.progress.setFormat('Отмена…')
            stopped = True
        if self._worker is not None:
            self._worker.stop()
            stopped = True
        if stopped:
            self.cancel_btn.setEnabled(False)

    def _set_busy(self, busy):
        self.progress.setVisible(busy)
        self.cancel_btn.setVisible(busy)
        self.cancel_btn.setEnabled(busy)
        self.open_btn.setEnabled(not busy)
        self.open_file_btn.setEnabled(not busy)
        self.recalc_btn.setEnabled(not busy and self._result is not None)
        for w in (self.beam_combo, self.freq_combo, self.beam_prev_btn,
                  self.beam_next_btn, self.freq_prev_btn, self.freq_next_btn,
                  self.order_combo):
            w.setEnabled(not busy)

    # --------------------------------------------------------- Навигация
    def _on_selection_changed(self, _index=None):
        if self._updating:
            return
        self._display_current()

    def _on_order_changed(self, _index=None):
        """Сменить порядок лучей в списке. Выбранный луч остаётся выбранным:
        _populate_combos ищет его по значению, поэтому перерисовка не нужна."""
        if self._updating:
            return
        self._beam_order = self.order_combo.currentData() or BEAM_ORDER_NUMBER
        self._settings().setValue('far_field/beam_order', self._beam_order)
        self._populate_combos()

    def _beam_prev(self):
        self._step_combo(self.beam_combo, -1)

    def _beam_next(self):
        self._step_combo(self.beam_combo, +1)

    def _freq_prev(self):
        self._step_combo(self.freq_combo, -1)

    def _freq_next(self):
        self._step_combo(self.freq_combo, +1)

    @staticmethod
    def _step_combo(combo, delta):
        new_index = combo.currentIndex() + delta
        if 0 <= new_index < combo.count():
            combo.setCurrentIndex(new_index)

    def _current_beam_freq(self):
        b_idx = self.beam_combo.currentIndex()
        f_idx = self.freq_combo.currentIndex()
        if not (0 <= b_idx < len(self._view_beams)) or not (0 <= f_idx < len(self._view_freqs)):
            return None, None
        return self._view_beams[b_idx], self._view_freqs[f_idx]

    # ------------------------------------------------------------ Вид окна
    def _set_view(self, view, force=False):
        """Переключить вид: 'near' — ближнее поле, 'far' — главные сечения ДН.

        В дальнюю зону пускаем только когда есть расчёт, иначе показывать нечего.
        """
        view = 'far' if view == 'far' else 'near'
        if view == 'far' and not self._cache:
            view = 'near'
        if view == self._view and not force:
            # Повторный клик по уже выбранной кнопке не должен её «отжимать».
            self.near_btn.setChecked(view == 'near')
            self.far_btn.setChecked(view == 'far')
            return
        self._view = view
        self.near_btn.setChecked(view == 'near')
        self.far_btn.setChecked(view == 'far')
        self.stack.setCurrentIndex(1 if view == 'far' else 0)
        self._sync_view_widgets()
        self._populate_combos()
        self._display_current()

    def _sync_view_widgets(self):
        """Показать то, что относится к текущему виду, и скрыть чужое."""
        far = self._view == 'far'
        computed = bool(self._cache)
        self.near_group.setVisible(not far)
        self.scan_group.setVisible(not far and self.scan_form.rowCount() > 0)
        self.metrics_group.setVisible(far)
        self.mask_group.setVisible(far)
        self.hold_btn.setEnabled(far and computed)
        self.clear_overlays_btn.setEnabled(far and computed)
        self.export_table_btn.setEnabled(far and computed)

    def _populate_combos(self):
        """Списки луча/частоты под текущий вид.

        В ближнем поле доступно всё загруженное, в дальней зоне — только
        рассчитанное (пересчёт часто идёт по части лучей).
        """
        prev_beam, prev_freq = self._current_beam_freq()
        if self._view == 'far':
            beams, freqs = list(self._computed_beams), list(self._computed_freqs)
        else:
            beams, freqs = list(self._beams), list(self._freqs)
        beams = order_beams(beams, self._beam_order)

        self._updating = True
        try:
            self._view_beams, self._view_freqs = beams, freqs
            self.beam_combo.clear()
            self.beam_combo.addItems([self._beam_item_text(b) for b in beams])
            self.freq_combo.clear()
            self.freq_combo.addItems([f'{f:g}' for f in freqs])
            self.beam_combo.setCurrentIndex(self._index_of(beams, prev_beam))
            self.freq_combo.setCurrentIndex(self._index_of(freqs, prev_freq))
        finally:
            self._updating = False

    @staticmethod
    def _index_of(values, previous):
        """Индекс previous в values, иначе первый элемент.

        Выбор сохраняется ПО ЗНАЧЕНИЮ: списки в видах разной длины, и по
        индексу после переключения оказался бы не тот луч.
        """
        if previous is not None:
            for i, value in enumerate(values):
                if value == previous:
                    return i
        return 0 if values else -1

    # ------------------------------------------------------- Луч и его углы
    def _beam_angles(self, beam):
        """(α, β) луча в градусах или None (в режиме файла луча нет)."""
        if self._single_file_mode or beam is None:
            return None
        try:
            mapping = beam_to_angles(int(beam))
            return float(mapping.alpha), float(mapping.beta)
        except Exception:
            return None

    def _beam_item_text(self, beam):
        """Номер луча вместе с углами отклонения — для списка выбора.

        В сгруппированном порядке ведущий угол идёт первым: так на глаз видно,
        где кончается одна группа и начинается следующая.
        """
        angles = self._beam_angles(beam)
        if angles is None:
            return str(beam)
        alpha, beta = angles
        if self._beam_order == BEAM_ORDER_AZIMUTH:
            return f'α={alpha:.2f}  ·  β={beta:.2f}  ·  луч {beam}'
        if self._beam_order == BEAM_ORDER_ELEVATION:
            return f'β={beta:.2f}  ·  α={alpha:.2f}  ·  луч {beam}'
        return f'{beam}  ·  α={alpha:.2f} β={beta:.2f}'

    def _current_label(self):
        beam, freq = self._current_beam_freq()
        if beam is None:
            return ''
        prefix = '' if self._single_file_mode else 'Луч '
        angles = self._beam_angles(beam)
        suffix = '' if angles is None else f' (α={angles[0]:.2f}° β={angles[1]:.2f}°)'
        return f'{prefix}{beam}{suffix} / {freq:g} МГц'

    # ----------------------------------------------------------- Отрисовка
    def _display_current(self):
        if self._view == 'near':
            self._display_near()
        else:
            self._display_far()

    def _near_step(self):
        """Шаг сканера для осей карты, см: из параметров, иначе из скана."""
        params = self._params or {}
        try:
            return (float(params.get('dx', self._default_step[0])),
                    float(params.get('dy', self._default_step[1])))
        except (TypeError, ValueError):
            return 1.0, 1.0

    def _display_near(self):
        """Карта ближнего поля выбранного луча/частоты."""
        beam, freq = self._current_beam_freq()
        field = None
        if beam is not None and self._result:
            field = (self._result.get('data', {}).get(beam) or {}).get(freq)
        if not field:
            self.near_panel.clear_data()
            self._update_near_metrics(None)
            self._update_scan_info(beam)
            return
        dx, dy = self._near_step()
        stats = self.near_panel.set_field(
            field['amp'], field['phase'],
            self._result.get('x_list'), self._result.get('y_list'),
            dx, dy, label=self._current_label())
        self._update_near_metrics(stats)
        self._update_scan_info(beam)

    # --------------------------------------------------- Сведения о скане
    def _beam_file(self, beam):
        """Файл текущего луча в папке скана или None, если его нет."""
        if self._single_file_mode:
            path = (self._result or {}).get('file_path')
            return path if path and os.path.isfile(path) else None
        folder = self._result.get('save_dir') or self._folder
        if not folder or beam is None:
            return None
        path = os.path.join(str(folder), f'Beam№{beam}.xlsx')
        return path if os.path.isfile(path) else None

    def _scan_info_rows(self, beam):
        """Пары (подпись, значение) о скане — только те, что есть в данных."""
        result = self._result or {}
        rows = []

        path = self._beam_file(beam)
        if path:
            rows.append(('Файл:', os.path.basename(path)))

        def span(lo_key, hi_key):
            lo, hi = result.get(lo_key), result.get(hi_key)
            if lo is None or hi is None:
                return None
            return f'{float(lo):g} … {float(hi):g}'

        x_span = span('left_x', 'right_x')
        if x_span:
            rows.append(('Область X, см:', x_span))
        y_span = span('up_y', 'down_y')
        if y_span:
            rows.append(('Область Y, см:', y_span))

        if result.get('step_x') is not None and result.get('step_y') is not None:
            rows.append(('Шаг X/Y, см:',
                         f"{float(result['step_x']):g} / {float(result['step_y']):g}"))

        freqs = result.get('freq_list') or []
        if freqs:
            rows.append(('Частоты, МГц:', ', '.join(f'{float(f):g}' for f in freqs)))

        rows += self._pna_info_rows(result.get('pna_settings'))
        rows += self._sync_info_rows(result.get('sync_settings'))
        return rows

    @staticmethod
    def _pna_info_rows(pna):
        """Строки о настройках PNA (частоты в Гц, импульсы в секундах)."""
        if not isinstance(pna, dict) or not pna:
            return []
        rows = []
        head = []
        if pna.get('s_param'):
            head.append(str(pna['s_param']))
        if pna.get('power') is not None:
            head.append(f"{float(pna['power']):g} дБм")
        if head:
            rows.append(('PNA:', ', '.join(head)))

        start, stop = pna.get('freq_start'), pna.get('freq_stop')
        if start is not None and stop is not None:
            text = f'{float(start) / 1e6:g} … {float(stop) / 1e6:g} МГц'
            if pna.get('freq_points'):
                text += f", точек: {int(pna['freq_points'])}"
            rows.append(('Диапазон PNA:', text))

        if pna.get('pulse_mode'):
            pulse = [str(pna['pulse_mode'])]
            if pna.get('pulse_period') is not None:
                pulse.append(f"период {float(pna['pulse_period']) * 1e6:g} мкс")
            if pna.get('pulse_width') is not None:
                pulse.append(f"ширина {float(pna['pulse_width']) * 1e6:g} мкс")
            rows.append(('Импульс:', ', '.join(pulse)))
        return rows

    @staticmethod
    def _sync_info_rows(sync):
        """Строки о синхронизаторе (времена хранятся в секундах)."""
        if not isinstance(sync, dict) or not sync:
            return []
        parts = []
        if sync.get('trig_ttl_channel'):
            parts.append(f"канал {sync['trig_ttl_channel']}")
        if sync.get('trig_start_lead') is not None:
            parts.append(f"опережение {float(sync['trig_start_lead']) * 1e3:g} мс")
        if sync.get('trig_pulse_period') is not None:
            parts.append(f"период {float(sync['trig_pulse_period']) * 1e6:g} мкс")
        return [('Синхронизатор:', ', '.join(parts))] if parts else []

    def _update_scan_info(self, beam=None):
        """Перестроить блок «Параметры скана» под текущие данные."""
        rows = self._scan_info_rows(beam) if self._result else []
        while self.scan_form.rowCount():
            self.scan_form.removeRow(0)
        for title, value in rows:
            lbl = self._metric()
            lbl.setText(value)
            lbl.setWordWrap(True)   # панель узкая, длинные значения переносим
            self.scan_form.addRow(title, lbl)
        # Пустую группу не показываем: у одиночного файла сведений может не быть.
        self.scan_group.setVisible(bool(rows) and self._view == 'near')

    def _update_near_metrics(self, stats):
        labels = (self.lbl_nf_angles, self.lbl_nf_max, self.lbl_nf_max_pos,
                  self.lbl_nf_dynamic, self.lbl_nf_phase_span,
                  self.lbl_nf_points, self.lbl_nf_size, self.lbl_nf_grid)
        if not stats:
            for lbl in labels:
                lbl.setText('—')
            return

        beam, _freq = self._current_beam_freq()
        angles = self._beam_angles(beam)
        self.lbl_nf_angles.setText(
            '—' if angles is None else f'{angles[0]:.2f} / {angles[1]:.2f}')

        def num(value, fmt='{:.2f}'):
            return '—' if value is None else fmt.format(value)

        self.lbl_nf_max.setText(num(stats['max_db']))
        if stats['max_x'] is None or stats['max_y'] is None:
            self.lbl_nf_max_pos.setText('—')
        else:
            self.lbl_nf_max_pos.setText(f"{stats['max_x']:.1f} / {stats['max_y']:.1f}")
        self.lbl_nf_dynamic.setText(num(stats['dynamic_db'], '{:.1f}'))
        self.lbl_nf_phase_span.setText(num(stats['phase_span'], '{:.1f}'))
        self.lbl_nf_points.setText(f"{stats['measured']} из {stats['total']}")
        if stats['size_x'] is None or stats['size_y'] is None:
            self.lbl_nf_size.setText('—')
        else:
            self.lbl_nf_size.setText(f"{stats['size_x']:.1f} × {stats['size_y']:.1f}")
        self.lbl_nf_grid.setText(f"{stats['n_x']} × {stats['n_y']}")

    def _display_far(self):
        beam, freq = self._current_beam_freq()
        if beam is None or freq is None:
            return
        entry = self._cache.get((beam, freq))
        if entry is None:
            self.plot_panel.clear_data()
            for lbl in (self.lbl_max, self.lbl_pos_az, self.lbl_pos_el, self.lbl_sll_az,
                        self.lbl_sll_el, self.lbl_bw_az, self.lbl_bw_el,
                        self.lbl_phase_max, self.lbl_set_angle):
                lbl.setText('—')
            self._update_mask()
            return
        self.plot_panel.set_data([
            (self._az_deg, entry['az_amp'], entry['max_val']),
            (self._el_deg, entry['el_amp'], entry['max_val']),
            (self._az_deg, entry['az_phase'], 0.0),
            (self._el_deg, entry['el_phase'], 0.0),
        ])
        self._update_metrics(beam, entry)
        self._update_mask()

    def _update_mask(self, *_):
        """Маска УБЛ: линия предела на графике + вердикт годен/не годен."""
        if not self.mask_check.isChecked():
            self.plot_panel.set_sll_mask(None)
            self.mask_result.setText('—')
            self.mask_result.setStyleSheet('')
            return
        db = self.mask_spin.value()
        self.plot_panel.set_sll_mask(db)
        beam, freq = self._current_beam_freq()
        entry = self._cache.get((beam, freq)) if beam is not None else None
        if entry is None:
            self.mask_result.setText('нет данных')
            self.mask_result.setStyleSheet('color:#667085;')
            return
        worst = None
        for arr in (entry['az_amp'], entry['el_amp']):
            sll = side_lobe_level(arr)
            if sll is not None:
                worst = sll if worst is None else max(worst, sll)
        if worst is None:
            self.mask_result.setText('боковых нет — годен')
            self.mask_result.setStyleSheet(f'color:{STATUS_ICON["ok"]}; font-weight:600;')
            return
        ok = worst <= db
        color = STATUS_ICON['ok'] if ok else STATUS_ICON['fail']
        verdict = 'ГОДЕН' if ok else 'НЕ ГОДЕН'
        self.mask_result.setText(f'{verdict}\nхудший УБЛ {worst:.1f} дБ (предел {db:.0f})')
        self.mask_result.setStyleSheet(f'color:{color}; font-weight:600;')

    @staticmethod
    def _beamwidth(x, amp):
        """Ширина главного лепестка по уровню −3 дБ (линейная интерполяция краёв)."""
        x = np.asarray(x, dtype=float)
        amp = np.asarray(amp, dtype=float)
        if amp.size < 3 or not np.any(np.isfinite(amp)):
            return None
        peak = int(np.nanargmax(amp))
        level = amp[peak] - 3.0

        def edge(rng):
            prev = peak
            for i in rng:
                if (amp[i] - level) * (amp[prev] - level) <= 0 and amp[i] != amp[prev]:
                    t = (level - amp[prev]) / (amp[i] - amp[prev])
                    return x[prev] + t * (x[i] - x[prev])
                prev = i
            return None

        left = edge(range(peak - 1, -1, -1))
        right = edge(range(peak + 1, amp.size))
        if left is None or right is None:
            return None
        return abs(right - left)

    def _update_metrics(self, beam, entry):
        self.lbl_max.setText(f"{entry['max_val']:.2f}")
        self.lbl_pos_az.setText(f"{entry['pos_az']:.3f}")
        self.lbl_pos_el.setText(f"{entry['pos_el']:.3f}")
        bw_az = self._beamwidth(self._az_deg, entry['az_amp'])
        bw_el = self._beamwidth(self._el_deg, entry['el_amp'])
        self.lbl_bw_az.setText('—' if bw_az is None else f"{bw_az:.3f}")
        self.lbl_bw_el.setText('—' if bw_el is None else f"{bw_el:.3f}")
        self.lbl_sll_az.setText('—' if entry['sll_az'] is None else f"{entry['sll_az']:.2f}")
        self.lbl_sll_el.setText('—' if entry['sll_el'] is None else f"{entry['sll_el']:.2f}")
        self.lbl_phase_max.setText(f"{entry['phase_max']:.2f}")
        try:
            mapping = beam_to_angles(int(beam))
            # Без префиксов «α=/β=»: их называет подпись строки, а панель узкая
            # (300 px) — с префиксами β не влезала и обрезалась.
            self.lbl_set_angle.setText(f'{mapping.alpha:.2f} / {mapping.beta:.2f}')
        except Exception:
            self.lbl_set_angle.setText('—')

    # ----------------------------------------------------------- Наложение
    def _hold_traces(self):
        label = self._current_label()
        if not label:
            return
        for p in self.panels:
            p.hold_current(label)

    def _clear_overlays(self):
        for p in self.panels:
            p.clear_overlays()

    # ------------------------------------------------ Пакетный экспорт метрик
    _METRICS_HEADERS = [
        'Луч', 'Частота, МГц', 'Максимум, дБ', 'Положение Az, °', 'Положение El, °',
        'Ширина Az (−3дБ), °', 'Ширина El (−3дБ), °', 'УБЛ Az, дБ', 'УБЛ El, дБ',
        'Фаза в макс., °', 'Задан. α, °', 'Задан. β, °',
    ]

    def _metrics_rows(self):
        """Строки таблицы метрик по всем рассчитанным (луч, частота).

        Лучи идут в том же порядке, что выбран в списке: таблица читается так
        же, как они перебираются в окне.
        """
        rows = []
        beams = order_beams(sorted({b for b, _ in self._cache}), self._beam_order)
        freqs = sorted({f for _, f in self._cache})
        for beam, freq in ((b, f) for b in beams for f in freqs):
            e = self._cache.get((beam, freq))
            if e is None:
                continue
            bw_az = self._beamwidth(self._az_deg, e['az_amp'])
            bw_el = self._beamwidth(self._el_deg, e['el_amp'])
            try:
                mp = beam_to_angles(int(beam))
                alpha, beta = mp.alpha, mp.beta
            except Exception:
                alpha = beta = None
            rows.append([
                beam, freq, round(e['max_val'], 2),
                round(e['pos_az'], 3), round(e['pos_el'], 3),
                None if bw_az is None else round(bw_az, 3),
                None if bw_el is None else round(bw_el, 3),
                None if e['sll_az'] is None else round(e['sll_az'], 2),
                None if e['sll_el'] is None else round(e['sll_el'], 2),
                round(e['phase_max'], 2), alpha, beta,
            ])
        return rows

    def _export_metrics_table(self):
        if not self._cache:
            return
        path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, 'Сохранить таблицу метрик', 'far_field_metrics.xlsx',
            'Excel (*.xlsx);;CSV (*.csv)')
        if not path:
            return
        rows = self._metrics_rows()
        try:
            if path.lower().endswith('.csv'):
                self._write_metrics_csv(path, self._METRICS_HEADERS, rows)
            else:
                self._write_metrics_xlsx(path, self._METRICS_HEADERS, rows)
            logger.info(f'Таблица метрик сохранена: {path} ({len(rows)} строк)')
            QtWidgets.QMessageBox.information(self, 'Готово',
                                             f'Сохранено строк: {len(rows)}\n{path}')
        except Exception as exc:
            logger.error(f'Ошибка экспорта таблицы метрик: {exc}', exc_info=True)
            QtWidgets.QMessageBox.critical(self, 'Ошибка экспорта', str(exc))

    @staticmethod
    def _write_metrics_csv(path, headers, rows):
        with open(path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f, delimiter=';')
            writer.writerow(headers)
            for r in rows:
                writer.writerow(['' if v is None else v for v in r])

    @staticmethod
    def _write_metrics_xlsx(path, headers, rows):
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
        wb = Workbook()
        ws = wb.active
        ws.title = 'Метрики'
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for r in rows:
            ws.append(r)
        for i, h in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(i)].width = max(12, len(h) + 2)
        ws.freeze_panes = 'A2'
        wb.save(path)

    # ------------------------------------------------------------- Закрытие
    def _close_if_pending(self):
        """Закрыть окно, если оно ждало остановки потоков (см. closeEvent)."""
        if self._closing and self._thread is None and self._load_thread is None:
            self.close()

    def closeEvent(self, event):
        for worker in (self._worker, self._load_worker):
            if worker is not None:
                worker.stop()
        threads = [t for t in (self._thread, self._load_thread) if t is not None]
        for thread in threads:
            thread.quit()
            thread.wait(2000)
        if any(t.isRunning() for t in threads):
            # Разбор книги Excel обрывается не мгновенно. Пока поток жив, окно не
            # закрываем: иначе Qt роняет приложение («QThread: Destroyed while
            # thread is still running»). Закроемся сами, как только поток встанет.
            self._closing = True
            self.progress.setVisible(True)
            self.progress.setRange(0, 0)
            self.progress.setFormat('Завершение работы…')
            event.ignore()
            return
        self._save_window_state()
        super().closeEvent(event)
